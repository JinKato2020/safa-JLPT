# 読解/聴解/漢字/語彙/文法 を Excel(.xlsx) に整理出力。再実行で常に最新JSONを反映。
# 出力: ../コンテンツ一覧.xlsx (プロジェクト直下)
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(ROOT, "..", "app", "src", "data")
RJSON = os.path.join(DATA, "reading.json")
LJSON = os.path.join(DATA, "listening.json")
KJSON = os.path.join(DATA, "kanji.json")
VJSON = os.path.join(DATA, "vocab.json")
GJSON = os.path.join(DATA, "grammar.json")
OUT = os.path.join(ROOT, "..", "コンテンツ一覧.xlsx")
LEVELS = ["N5", "N4", "N3"]

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="2563EB")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_sheet(wb, title, passages, passage_cols):
    ws = wb.create_sheet(title)
    if passage_cols == "reading":
        head = ["級", "ID", "形式", "タイトル", "本文", "設問ID", "設問", "選択肢1", "選択肢2", "選択肢3", "選択肢4", "正答番号", "正答", "解説"]
        w = [6, 10, 10, 22, 50, 12, 30, 18, 18, 18, 18, 8, 22, 34]
    else:
        head = ["級", "ID", "タイトル", "スクリプト", "設問ID", "設問", "選択肢1", "選択肢2", "選択肢3", "選択肢4", "正答番号", "正答", "解説"]
        w = [6, 10, 22, 50, 12, 30, 18, 18, 18, 18, 8, 22, 34]
    ws.append(head)
    style_header(ws, len(head))
    widths(ws, w)
    order = {lv: i for i, lv in enumerate(LEVELS)}
    for p in sorted(passages, key=lambda x: (order.get(x["level"], 9), x["id"])):
        first = True
        for q in p["questions"]:
            ai = q["answerIndex"]
            ch = (q["choices"] + ["", "", "", ""])[:4]
            if passage_cols == "reading":
                row = [p["level"] if first else "", p["id"] if first else "", p.get("format", "") if first else "",
                       p.get("title", "") if first else "", p.get("body", "") if first else "",
                       q["id"], q["q"], ch[0], ch[1], ch[2], ch[3], ai + 1, ch[ai], q.get("explain", "")]
            else:
                row = [p["level"] if first else "", p["id"] if first else "", p.get("title", "") if first else "",
                       p.get("script", "") if first else "",
                       q["id"], q["q"], ch[0], ch[1], ch[2], ch[3], ai + 1, ch[ai], q.get("explain", "")]
            ws.append(row)
            for c in range(1, len(head) + 1):
                ws.cell(row=ws.max_row, column=c).alignment = WRAP if c in (5, 7, 14, 4, 6, 13) else TOP
            first = False
    return ws


def add_flat_sheet(wb, title, items, cols):
    """漢字/語彙/文法のようなフラットなリストをシート化。cols=(見出し, key or fn, 幅, 折返し)"""
    ws = wb.create_sheet(title)
    ws.append([c[0] for c in cols])
    style_header(ws, len(cols))
    widths(ws, [c[2] for c in cols])
    order = {lv: i for i, lv in enumerate(LEVELS)}
    for it in sorted(items, key=lambda x: (order.get(x["level"], 9), x["id"])):
        row = [(getter(it) if callable(getter) else it.get(getter, "")) for (_, getter, _, _) in cols]
        ws.append(row)
        for ci, (_, _, _, wrap) in enumerate(cols, 1):
            ws.cell(row=ws.max_row, column=ci).alignment = WRAP if wrap else TOP
    return ws


def overview(wb, reading, listening, kanji, vocab, grammar):
    ws = wb.create_sheet("概要", 0)
    ws.append(["級", "読解 本数", "読解 設問数", "聴解 本数", "聴解 設問数", "漢字", "語彙", "文法"])
    style_header(ws, 8)
    widths(ws, [8, 11, 11, 11, 11, 8, 8, 8])

    def cnt(arr, lv):
        return len([x for x in arr if x["level"] == lv])

    for lv in LEVELS:
        rp = [p for p in reading if p["level"] == lv]
        lp = [p for p in listening if p["level"] == lv]
        ws.append([lv, len(rp), sum(len(p["questions"]) for p in rp),
                   len(lp), sum(len(p["questions"]) for p in lp),
                   cnt(kanji, lv), cnt(vocab, lv), cnt(grammar, lv)])
    ws.append(["合計", len(reading), sum(len(p["questions"]) for p in reading),
               len(listening), sum(len(p["questions"]) for p in listening),
               len(kanji), len(vocab), len(grammar)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    ws.append(["※「本数」=文章/会話クリップの数、「設問数」=その問題(4択)の総数。漢字/語彙/文法は項目数。"])


KANJI_COLS = [("級", "level", 6, False), ("ID", "id", 10, False), ("漢字", "char", 8, False),
              ("音読み", "on", 18, True), ("訓読み", "kun", 22, True), ("意味", "meaning", 32, True),
              ("画数", "strokes", 8, False), ("学年", "grade", 8, False)]
VOCAB_COLS = [("級", "level", 6, False), ("ID", "id", 10, False), ("語", "word", 20, True),
              ("読み", "reading", 20, True), ("意味", "meaning", 40, True),
              ("タグ", lambda x: "、".join(x.get("tags") or []), 24, True)]
GRAMMAR_COLS = [("級", "level", 6, False), ("ID", "id", 10, False), ("文法項目", "point", 28, True),
                ("ローマ字", "romaji", 24, True), ("意味", "meaning", 32, True),
                ("例文", "exampleJa", 42, True), ("訳", "exampleEn", 42, True)]


def main():
    reading = json.load(open(RJSON, encoding="utf-8"))
    listening = json.load(open(LJSON, encoding="utf-8"))
    kanji = json.load(open(KJSON, encoding="utf-8"))
    vocab = json.load(open(VJSON, encoding="utf-8"))
    grammar = json.load(open(GJSON, encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)
    overview(wb, reading, listening, kanji, vocab, grammar)
    add_sheet(wb, "読解", reading, "reading")
    add_sheet(wb, "聴解", listening, "listening")
    add_flat_sheet(wb, "漢字", kanji, KANJI_COLS)
    add_flat_sheet(wb, "語彙", vocab, VOCAB_COLS)
    add_flat_sheet(wb, "文法", grammar, GRAMMAR_COLS)
    wb.save(OUT)
    print("saved ->", os.path.abspath(OUT))
    print(f"読解{len(reading)}本/{sum(len(p['questions']) for p in reading)}問  "
          f"聴解{len(listening)}本/{sum(len(p['questions']) for p in listening)}問  "
          f"漢字{len(kanji)}  語彙{len(vocab)}  文法{len(grammar)}")


if __name__ == "__main__":
    main()
