# safa JLPT — データパイプライン: JLPT_N5-N3_完全データ.xlsx → app/src/data/*.json
# 再現可能なビルド資産。実行: python data-build/build_data.py
# 出典: 漢字=KANJIDIC / 語彙=JMdict (© EDRDG, CC BY-SA)。文法D列・例文・活用は safa オリジナル。
import openpyxl, json, os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'JLPT_N5-N3_完全データ.xlsx')
OUT = os.path.join(ROOT, 'app', 'src', 'data')
os.makedirs(OUT, exist_ok=True)

LEVELS = ['N5', 'N4', 'N3']
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def body_rows(ws):
    """ヘッダ(1行目)を除く本文行。#列(0)が空なら終端とみなしスキップ。"""
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or r[0] is None:
            continue
        yield list(r) + [None] * 8  # 末尾不足をNone埋め


def s(x):
    return x.strip() if isinstance(x, str) else x


# ソースデータの既知の誤りを補正(語→正しい語/読み/意味)。例: ラジオカセ=ラジオカセットの途中切れ→ラジカセ。
VOCAB_FIX = {
    'ラジオカセ': {'word': 'ラジカセ', 'reading': 'ラジカセ', 'meaning': 'boombox; radio-cassette player'},
}


def fix_vocab(word, reading, meaning):
    f = VOCAB_FIX.get(word)
    if f:
        return f['word'], f['reading'], f.get('meaning', meaning)
    return word, reading, meaning


def clean_reading(word, reading):
    """xlsxソースに紛れる『する動詞』混入(運動→うんどうする 等)を補正。
    語が する で終わらないのに読みが する で終わる場合は除去(辞書/読み問題の誤読み防止)。"""
    if isinstance(word, str) and isinstance(reading, str) \
            and reading.endswith('する') and not word.endswith('する'):
        return reading[:-2]
    return reading


kanji, vocab, grammar = [], [], []

for lv in LEVELS:
    lp = lv.lower()
    # 漢字: #, 漢字, 音読み, 訓読み, 意味(英), 画数, 学年
    for r in body_rows(wb[f'{lv}_漢字']):
        kanji.append({
            'id': f'{lp}-k-{int(r[0])}', 'level': lv,
            'category': 'moji_goi', 'type': 'kanji',
            'char': s(r[1]), 'on': s(r[2]), 'kun': s(r[3]),
            'meaning': s(r[4]), 'strokes': r[5], 'grade': r[6],
        })
    # 語彙: #, 語, 読み, 意味(英), タグ(旧/新JLPT)
    for r in body_rows(wb[f'{lv}_語彙']):
        tags = r[4].split() if isinstance(r[4], str) else []
        w, rd, mn = fix_vocab(s(r[1]), clean_reading(s(r[1]), s(r[2])), s(r[3]))
        vocab.append({
            'id': f'{lp}-v-{int(r[0])}', 'level': lv,
            'category': 'moji_goi', 'type': 'vocab',
            'word': w, 'reading': rd, 'meaning': mn,
            'tags': tags,
        })
    # 文法: #, 文法(日本語), 読み/ローマ字, 意味・用法(英), 例文(日本語ふりがな), 例文の英訳
    for r in body_rows(wb[f'{lv}_文法']):
        grammar.append({
            'id': f'{lp}-g-{int(r[0])}', 'level': lv,
            'category': 'bunpou', 'type': 'grammar',
            'point': s(r[1]), 'romaji': s(r[2]), 'meaning': s(r[3]),
            'exampleJa': s(r[4]), 'exampleEn': s(r[5]),
        })

# 動詞活用: 1行目ヘッダ(改行入り)→ 先頭行をキーに
verbs = []
ws = wb['動詞活用']
header = None
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        header = [(c.split('\n')[0].strip() if isinstance(c, str) else c) for c in r]
        continue
    if r is None or r[0] is None:
        continue
    verbs.append({k: s(v) for k, v in zip(header, r) if k})

# メタ: 公式配点(掲示板§5・出典 jlpt.jp)。合格=総合点＋各区分基準点の両方。
meta = {
    'levels': LEVELS,
    # UI 4リングの内訳(文字語彙=漢字+語彙)。公式区分はこれより粗い(下記 sections)。
    'ringCategories': ['moji_goi', 'bunpou', 'dokkai', 'choukai'],
    'ringLabels': {'moji_goi': '文字語彙', 'bunpou': '文法', 'dokkai': '読解', 'choukai': '聴解'},
    # 公式区分と基準点。N4/N5 は「言語知識・読解」が合算(120点満点・基準38)、聴解(60・基準19)。
    # N3 は 言語知識(60)/読解(60)/聴解(60) の3区分・各基準19。満点は全級180。
    'passMarks': {
        'N5': {'overall': 80, 'maxTotal': 180, 'timeMin': 90,
               'sections': {'gengo_dokkai': {'max': 120, 'min': 38}, 'choukai': {'max': 60, 'min': 19}}},
        'N4': {'overall': 90, 'maxTotal': 180, 'timeMin': 115,
               'sections': {'gengo_dokkai': {'max': 120, 'min': 38}, 'choukai': {'max': 60, 'min': 19}}},
        'N3': {'overall': 95, 'maxTotal': 180, 'timeMin': 140,
               'sections': {'gengo': {'max': 60, 'min': 19}, 'dokkai': {'max': 60, 'min': 19}, 'choukai': {'max': 60, 'min': 19}}},
    },
    'license': '辞書データ: 漢字 KANJIDIC / 語彙 JMdict (© EDRDG, CC BY-SA). 文法・例文・活用は safa オリジナル.',
}


def dump(name, obj):
    with open(os.path.join(OUT, name), 'w', encoding='utf-8') as f:
        json.dump(obj, f, ensure_ascii=False, separators=(',', ':'))
    return len(obj) if isinstance(obj, list) else 1


counts = {
    'kanji.json': dump('kanji.json', kanji),
    'vocab.json': dump('vocab.json', vocab),
    'grammar.json': dump('grammar.json', grammar),
    'verbs.json': dump('verbs.json', verbs),
}
dump('meta.json', meta)

# 検証: 掲示板§7の期待件数と一致するか
EXPECT = {'kanji': 612, 'vocab': 3526, 'grammar': 394, 'verbs': 29}
got = {'kanji': len(kanji), 'vocab': len(vocab), 'grammar': len(grammar), 'verbs': len(verbs)}
print('counts:', got)
ok = all(got[k] == EXPECT[k] for k in EXPECT)
# 級別内訳
for lv in LEVELS:
    print(f'  {lv}: 漢字={sum(1 for x in kanji if x["level"]==lv)} '
          f'語彙={sum(1 for x in vocab if x["level"]==lv)} '
          f'文法={sum(1 for x in grammar if x["level"]==lv)}')
print('EXPECT match (612/3526/394/29):', 'OK' if ok else f'MISMATCH expected {EXPECT}')
print('output ->', OUT)
