# _exam_bank.jsonl → 読解・聴解 問題バンク Excel。レベル別シート＋集計。セッション直下に出力。
import json, os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
src = os.path.join(ROOT, 'data-build', '_exam_bank.jsonl')
items = [json.loads(l) for l in open(src, encoding='utf-8') if l.strip()]

LV_ORDER = {'N5': 0, 'N4': 1, 'N3': 2}
KU_ORDER = {'読解': 0, '聴解': 1}
DAI_ORDER = {'内容理解(短文)': 0, '内容理解(中文)': 1, '内容理解(長文)': 2, '情報検索': 3,
             '課題理解': 4, 'ポイント理解': 5, '概要理解': 6, '発話表現': 7, '即時応答': 8}
items.sort(key=lambda x: (LV_ORDER.get(x['level'], 9), KU_ORDER.get(x['section'], 9), DAI_ORDER.get(x['daimon'], 9)))

cols = ['区分', '大問', '問われる力', '本文・台本', '設問', '正解', '誤答1', '誤答2', '誤答3', '解説']
widths = [7, 16, 18, 52, 26, 22, 18, 18, 18, 30]
hdrFill = PatternFill('solid', fgColor='2563EB')
readFill = PatternFill('solid', fgColor='EAF1FF')
listenFill = PatternFill('solid', fgColor='FFF3E0')
thin = Side(style='thin', color='D9D9D9')

wb = Workbook()
wb.remove(wb.active)

# 集計シート
ws0 = wb.create_sheet('集計')
ws0.append(['レベル', '区分', '大問', '問題数'])
for i, w in enumerate([8, 8, 18, 8], 1):
    ws0.cell(1, i).font = Font(bold=True, color='FFFFFF'); ws0.cell(1, i).fill = hdrFill
    ws0.column_dimensions[chr(64 + i)].width = w
cnt = Counter((x['level'], x['section'], x['daimon']) for x in items)
for (lv, ku, dai), n in sorted(cnt.items(), key=lambda kv: (LV_ORDER.get(kv[0][0], 9), KU_ORDER.get(kv[0][1], 9), DAI_ORDER.get(kv[0][2], 9))):
    ws0.append([lv, ku, dai, n])
ws0.append(['合計', '', '', len(items)])
ws0.cell(ws0.max_row, 1).font = Font(bold=True); ws0.cell(ws0.max_row, 4).font = Font(bold=True)

# レベル別シート
for lv in ['N5', 'N4', 'N3']:
    rows = [x for x in items if x['level'] == lv]
    if not rows:
        continue
    ws = wb.create_sheet(f'{lv}（{len(rows)}問）')
    ws.append(cols)
    for i, w in enumerate(widths, 1):
        cell = ws.cell(1, i); cell.font = Font(bold=True, color='FFFFFF'); cell.fill = hdrFill
        cell.alignment = Alignment(vertical='center', horizontal='center'); ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 20
    for x in rows:
        ch = x.get('choices', ['', '', '', ''])
        ws.append([x['section'], x['daimon'], x.get('power', ''), x.get('stem', ''), x['question'],
                   ch[0] if len(ch) > 0 else '', ch[1] if len(ch) > 1 else '', ch[2] if len(ch) > 2 else '', ch[3] if len(ch) > 3 else '',
                   x.get('explain', '')])
        ri = ws.max_row
        fill = readFill if x['section'] == '読解' else listenFill
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(ri, ci); cell.alignment = Alignment(vertical='top', wrap_text=True); cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if ci in (1, 2):
                cell.fill = fill
        ws.cell(ri, 6).font = Font(bold=True, color='1B7A1B')
        ws.row_dimensions[ri].height = 70

out = os.path.join(ROOT, 'JLPT_読解聴解_問題バンク.xlsx')
wb.save(out)
print('保存:', out, '/', len(items), '問')
