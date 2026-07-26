# _compare.json → 案A/案B 比較 Excel。問題・正解・誤答・解説・検証を並べる。
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
data = json.load(open(os.path.join(ROOT, 'data-build', '_compare.json'), encoding='utf-8'))
# 比較しやすいよう 種別→案 で並べる(読解の案A→案B→聴解の案A→案B)
data.sort(key=lambda r: (r['kind'], r['tag']))

wb = Workbook(); ws = wb.active; ws.title = '案A vs 案B'
cols = ['案', '種別', 'タイトル', '本文・台本', '設問', '正解', '誤答1', '誤答2', '誤答3', '解説', '検証', '検証理由']
widths = [11, 7, 22, 50, 28, 22, 18, 18, 18, 30, 8, 30]
ws.append(cols)
hdrFill = PatternFill('solid', fgColor='2563EB'); thin = Side(style='thin', color='D0D0D0')
for i, (col, w) in enumerate(zip(cols, widths), 1):
    cell = ws.cell(1, i); cell.font = Font(bold=True, color='FFFFFF'); cell.fill = hdrFill
    cell.alignment = Alignment(vertical='center', horizontal='center'); ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 22

aFill = PatternFill('solid', fgColor='EAF1FF'); bFill = PatternFill('solid', fgColor='FFF6E6')
okFill = PatternFill('solid', fgColor='E6F6E6'); ngFill = PatternFill('solid', fgColor='FDE8E8')
for r in data:
    it = r['it']; q = it
    body = it.get('body') or it.get('script') or ''
    ch = it.get('choices', ['', '', '', ''])
    row = [r['tag'], r['kind'], it.get('title', ''), body, it.get('q', ''),
           ch[0] if len(ch) > 0 else '', ch[1] if len(ch) > 1 else '', ch[2] if len(ch) > 2 else '', ch[3] if len(ch) > 3 else '',
           it.get('explain', ''), '合格' if r['valid'] else '不合格', r.get('reason', '')]
    ws.append(row)
    ri = ws.max_row
    rowfill = aFill if '案A' in r['tag'] else bFill
    for ci in range(1, len(cols)+1):
        cell = ws.cell(ri, ci); cell.alignment = Alignment(vertical='top', wrap_text=True); cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if ci in (1, 2): cell.fill = rowfill
    ws.cell(ri, 6).font = Font(bold=True, color='1B7A1B')  # 正解=緑太字
    ws.cell(ri, 11).fill = okFill if r['valid'] else ngFill
    ws.row_dimensions[ri].height = 86

out = os.path.join(ROOT, '案A_案B_問題比較.xlsx')
wb.save(out)
print('保存:', out, '/', len(data), '件 (案A=4o / 案B=mini・読解→聴解の順)')
