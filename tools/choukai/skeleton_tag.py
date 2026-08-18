"""聴解 骨組みパラメータのタグ付け＋分布(census)＋番人(check)＝恒久ツール(2026-08-18)。
設計正本＝md/聴解_作問フロー.md「骨組みパラメータの最適化」。ワンパターン化を機械で止めるための作問メタ。

対象フィールド（content/problems/choukai/*.json の各item に付与）:
  ① 課題(kadai)  develop … くぐり方=答えの決まり方(8値)。※設問型は既存qtype_ledgerで別途点検
  ② ポイント(point) kanten … 観点(なぜ/いつ/いくつ/気持ち/どれ)。qtype_ledger.kantenを再利用=二重管理しない
  ③ 概要(gaiyou) genre … 話題分野(8値) ／ q_type … 聞き方(何について/主張/タイトル)

判定の分担:
  kanten・q_type = 決定的(regex)＝backfillで即付与(正確)
  develop・genre = 会話ロジック/話題の判断が要る＝LLM分類の結果を --apply-map <map.json> で流し込む
                   (map.json = {"<id>": "<値>", ...}。値は下記TAXONOMYのいずれか)

モード:
  backfill  … kanten(point)・q_type(gaiyou) を決定的に付与して書き戻し
  apply-map … develop/genre を {id:値} のmapから付与して書き戻し(値の妥当性を検査)
  census    … 各フィールドの現在の分布を表示(未付与=「(未)」で計上)
  check     … 番人。違反(欠落あり／最大シェア>35%／kanten「気持ち」=0)でexit 1
使い方:
  python tools/choukai/skeleton_tag.py backfill
  python tools/choukai/skeleton_tag.py apply-map develop map_develop.json
  python tools/choukai/skeleton_tag.py census
  python tools/choukai/skeleton_tag.py check
"""
import json, os, sys, collections
from qtype_ledger import kanten as kanten_of, qtype as qtype_of, strip_furi

DIR = os.path.join(os.path.dirname(__file__), "..", "..", "content", "problems", "choukai")
KADAI = ["kadai_N5", "kadai_N4", "kadai_N3"]
POINT = ["point_N5", "point_N4", "point_N3"]
GAIYOU = ["gaiyou_N3"]

# ── TAXONOMY（設計＝md/聴解_作問フロー.md と一致。apply-mapの値検査に使う）──
DEVELOP = ["上書き", "条件順序", "消去", "まず次", "追加", "二者択一", "断って代案", "勘違い訂正"]
GENRE   = ["健康体", "生活くらし", "社会自然", "仕事", "学び", "モノサービス", "文化行事", "趣味旅食"]
QTYPE_G = ["何について", "主張", "タイトル"]

MAX_SHARE = 0.35  # 番人：1つの値がこれを超えたら偏り＝失敗


def load(name):
    p = os.path.join(DIR, name + ".json")
    with open(p, encoding="utf-8") as f:
        return p, json.load(f)


def save(p, d):
    with open(p, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
        f.write("\n")


def q_of(it):
    return it["questions"][0]["q"]


def gaiyou_qtype(q):
    q = strip_furi(q or "").replace(" ", "").replace("　", "")
    if "について" in q:
        return "何について"
    if ("タイトル" in q) or ("題名" in q) or ("見出し" in q):
        return "タイトル"
    # 一番言いたい/伝えたい/主に何を/最も言いたい 等＝主張
    return "主張"


def backfill():
    # ② point → kanten
    for name in POINT:
        p, d = load(name)
        for it in d["items"]:
            it["kanten"] = kanten_of(q_of(it))
        save(p, d)
    # ③ gaiyou → q_type
    for name in GAIYOU:
        p, d = load(name)
        for it in d["items"]:
            it["q_type"] = gaiyou_qtype(q_of(it))
        save(p, d)
    print("[backfill] kanten(point 3ファイル)・q_type(gaiyou) 付与済み")


def apply_map(field, mapfile):
    if field not in ("develop", "genre"):
        print("apply-map の field は develop か genre のみ"); sys.exit(2)
    allowed = set(DEVELOP if field == "develop" else GENRE)
    files = KADAI if field == "develop" else GAIYOU
    with open(mapfile, encoding="utf-8") as f:
        m = json.load(f)
    bad = sorted(set(m.values()) - allowed)
    if bad:
        print(f"⚠不正な値(TAXONOMY外)={bad}\n許可={sorted(allowed)}"); sys.exit(2)
    applied = miss = 0
    for name in files:
        p, d = load(name)
        for it in d["items"]:
            if it["id"] in m:
                it[field] = m[it["id"]]; applied += 1
            elif field not in it:
                miss += 1
        save(p, d)
    print(f"[apply-map {field}] 付与={applied}・未マップ残={miss}")


def census():
    def show(names, field, taxonomy):
        print(f"\n=== {field} ===")
        for name in names:
            _, d = load(name)
            c = collections.Counter(it.get(field, "(未)") for it in d["items"])
            n = len(d["items"])
            top = c.most_common(1)[0]
            miss = c.get("(未)", 0)
            order = taxonomy + [k for k in c if k not in taxonomy and k != "(未)"]
            dist = " ".join(f"{k}{c[k]}" for k in order if c.get(k))
            flag = "  ⚠偏り" if top[0] != "(未)" and top[1] / n > MAX_SHARE else ""
            print(f"  {name:11} n={n:3} 最大={top[0]}:{round(top[1]/n*100)}% 未付与={miss}{flag}")
            print(f"      {dist}")
    show(KADAI, "develop", DEVELOP)
    show(POINT, "kanten", ["なぜ", "いつ", "いくつ", "気持ち", "どれ"])
    show(GAIYOU, "genre", GENRE)
    show(GAIYOU, "q_type", QTYPE_G)


def check():
    # 作問前の助言ツール（厳しめ35%）。q_typeは音声固定で無償調整不可＝存在のみ確認しshareは対象外。
    problems = []
    present = [(KADAI, "develop"), (POINT, "kanten"), (GAIYOU, "genre"), (GAIYOU, "q_type")]
    share = [(KADAI, "develop"), (POINT, "kanten"), (GAIYOU, "genre")]
    for names, field in present:
        for name in names:
            _, d = load(name)
            miss = [it["id"] for it in d["items"] if not it.get(field)]
            if miss:
                problems.append(f"{name}.{field}: 未付与{len(miss)}件 (例:{miss[:3]})")
    for names, field in share:
        for name in names:
            _, d = load(name)
            items = d["items"]; n = len(items)
            c = collections.Counter(it.get(field) for it in items if it.get(field))
            if c:
                k, v = c.most_common(1)[0]
                if v / n > MAX_SHARE:
                    problems.append(f"{name}.{field}: 偏り『{k}』{round(v/n*100)}%>{int(MAX_SHARE*100)}%（作問で薄い値を優先）")
    # kanten「気持ち」は各レベル必須（公式に必ず出る）
    for name in POINT:
        _, d = load(name)
        if not any(it.get("kanten") == "気持ち" for it in d["items"]):
            problems.append(f"{name}.kanten: 「気持ち」観点が0（公式必須）")
    if problems:
        print("❌ 番人NG:")
        for p in problems:
            print("  -", p)
        sys.exit(1)
    print("✅ 番人OK（欠落なし・偏り≤35%・気持ち観点あり）")


KANTEN_ORDER = ["なぜ", "いつ", "いくつ", "気持ち", "どれ"]
# (大問ラベル, ファイル群, フィールド, 値集合, 音声固定で薄めは新規追加のみか)
XLSX_PLAN = [
    ("①課題理解", KADAI, "develop", DEVELOP, False),
    ("②ポイント理解", POINT, "kanten", KANTEN_ORDER, False),
    ("③概要理解", GAIYOU, "genre", GENRE, False),
    ("③概要理解", GAIYOU, "q_type", QTYPE_G, True),
]


def write_xlsx():
    """在庫・模試ストックまとめ.xlsx にシート「聴解 骨組みパラメータ分布」を追加/更新。
    「薄める対象＝どの大問・レベル・型を増やせばよいか」が一目で分かる。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    xlsx = os.path.join(os.path.dirname(__file__), "..", "..", "memory", "在庫・模試ストックまとめ.xlsx")
    wb = load_workbook(xlsx)
    sh = "聴解 骨組みパラメータ分布"
    if sh in wb.sheetnames:
        del wb[sh]
    ws = wb.create_sheet(sh)
    notes = [
        "聴解 骨組みパラメータ分布（2026-08-18 自動集計｜skeleton_tag.py）＝ワンパターン化を止めるための型バランス。",
        "偏り＝聞かず解ける/飽きるの芽。直し方＝既存はいじらず『薄い型』を新しく足して薄める（既存音声の焼き直し不要・新規追加分だけTTS）。",
        "番人：1つの型が過半(50%)を占めたらビルド失格。作問前の目安＝35%以下（skeleton_tag.py check）。『薄い型』列＝ここを増やせば偏りが下がる。",
        "kanten(観点)は全レベル良好。q_type(概要の聞き方)は質問が音声焼込み＝既存の付け替えは有料、薄めは新規追加で。",
    ]
    for tx in notes:
        ws.append([tx])
    cols = ["大問", "レベル", "パラメータ", "問題数", "型の数", "目安/型", "最も多い型", "最大%", "判定", "薄い型＝ここを増やして薄める", "全分布"]
    ws.append(cols)
    hr = len(notes) + 1
    for c in ws[hr]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDEBF7"); c.alignment = Alignment(wrap_text=True, vertical="center")
    for label, files, field, values, audio_locked in XLSX_PLAN:
        for name in files:
            _, d = load(name)
            items = d["items"]; n = len(items)
            c = collections.Counter(it.get(field, "(未)") for it in items)
            guide = max(1, round(n / len(values)))
            order = values + [k for k in c if k not in values and k != "(未)"]
            dist = " ".join(f"{k}{c[k]}" for k in order if c.get(k))
            top = max(values, key=lambda k: c.get(k, 0))
            tmax = round(c.get(top, 0) / n * 100)
            if tmax > 35:
                # 偏った行＝最多以外で平均未満の型すべて＝ここを増やせば薄まる
                thin = [f"{k}({c.get(k,0)})" for k in values if k != top and c.get(k, 0) < guide]
            else:
                # 良好な行＝ノイズを出さず、極端に薄い型(平均の半分未満)だけ変化ギャップとして表示
                thin = [f"{k}({c.get(k,0)})" for k in values if c.get(k, 0) < guide * 0.5]
            if tmax >= 50:
                verd = "❌過半"
            elif tmax > 35:
                verd = "⚠偏り→薄める"
            else:
                verd = "✅良好"
            if audio_locked:
                verd += "(音声固定)"
            lv = name.split("_")[-1]
            ws.append([label, lv, field, n, len(values), guide, top, f"{tmax}%", verd, " ".join(thin) or "—", dist])
    for i, w in enumerate([14, 6, 9, 7, 6, 7, 12, 7, 16, 34, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    wb.save(xlsx)
    print(f"書込み: シート「{sh}」 → {xlsx}")


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "census"
    if mode == "backfill":
        backfill()
    elif mode == "apply-map":
        apply_map(sys.argv[2], sys.argv[3])
    elif mode == "census":
        census()
    elif mode == "check":
        check()
    elif mode == "--xlsx":
        write_xlsx()
    else:
        print(__doc__)


if __name__ == "__main__":
    main()
