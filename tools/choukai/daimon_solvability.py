# -*- coding: utf-8 -*-
r"""聴解 全大問×全レベルの「攻略耐性・モーラ・ワンパターン」分析＝恒久ツール（2026-08-16）。

【なぜ】発話表現で入れた攻略耐性の考え方を他大問へ移植。本質は3つ：
  ① 答えが本文/設問から予測できる設計になっていないか（＝聞かなくても解ける）
  ② 誤答が弱くて正解が簡単に見つからないか
  ③ 大問×レベル内で問題が類似・重複＝ワンパターン化していないか
【測り方（機械・近似）】
  ・語彙マッチ的中% … 「本文と最も語が重なる選択肢」を選ぶと正解になる割合。基準=1/選択肢数。
       高い＝正解だけ本文の語を使い誤答が本文と無関係＝聞かず語マッチで解ける（①②）。良い設計は正解を“言い換え”、
       本文の語は誤答に罠として置く（＝語マッチ的中は基準付近に下がる）。
  ・正解−誤答 本文一致差 … 正解の本文一致語数 − 誤答平均。プラスに大きいほど「正解だけ本文語」（①②）。
  ・最長を選ぶ的中% … 正解が最長選択肢の割合。基準=1/選択肢数。長さの手がかり（②）。表示シャッフルでも文長は不変。
  ・台本重複 / 選択肢セット重複 … 同一大問×レベル内で近似（③ワンパターン）。
  ・設問テンプレ最大% … 同一設問文の最大シェア（③）。
  ・（発話のみ）依頼形%/形分離%/機能・軸偏在。（音声焼込み大問のみ）正解位置の偏り。

選択肢方式：課題/ポイント/概要=テキスト4択で**表示時シャッフル**（正解位置は無関係）／発話/即時=音声焼込み（位置固定）。

使い方:
  python tools/choukai/daimon_solvability.py                 # 全大問×全レベルを集計・print
  python tools/choukai/daimon_solvability.py --xlsx          # 上記＋在庫Excelへシート「聴解攻略耐性分析」を更新
  python tools/choukai/daimon_solvability.py --draft d.json kadai N3   # 作問ドラフト(items配列)をゲート採点（将来作問QA）
しきい値（新規作問の目安・既存は帯導入前で超過あり）:
  語彙マッチ ≤ 基準+10pt / 最長 ≤ 基準+15pt / 正解−誤答差 ≤ +1.0 / 台本・選択肢重複 = 0 / 設問テンプレ ≤ 30%
"""
import sys, os, re, json, glob, statistics, argparse
from collections import Counter
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.dirname(glob.glob(os.path.join(ROOT, '**', 'merge_and_gate.py'), recursive=True)[0]))
from merge_and_gate import body_mora, strip_furi, body_text
from hatsuwa_axes import _form, classify_axis, classify_function
from sokuji_sim import nearest
import sokuji_build as SB
import merge_and_gate as MG

BAND = MG.load_bands()
REQ = {"依頼謙", "依頼", "依頼砕", "ください", "希望前置"}
KAKARI = re.compile(r"(?<![関])係(?![長り])")
ACCENT = {"留守"}
TOK = re.compile(r'[一-龿々〆]+|[ァ-ヶー]{2,}')   # 内容語＝漢字連続 or カタカナ2+
DUP_SCRIPT, DUP_CHSET = 0.60, 0.70
# 列挙型の答え（数量・曜日・助数詞）＝別会話でも選択肢集合が一致するのは正当。選択肢セット重複から除外。
_NUM = r'(?:一|二|三|四|五|六|七|八|九|十|百|千|\d|ひと|ふた|みっ|よっ|いつ|むっ|なな|やっ|ここの|とお|いく)'
_CNT = r'(?:つ|回|枚|人|個|本|冊|時|分|時間|日|月|年|円|番|階|度|杯|匹|台|歳|才|ページ|キロ|メートル|グラム|割|%|パーセント|km|m)?'
_DAY = r'(?:月|火|水|木|金|土|日)曜日|(?:げつ|か|すい|もく|きん|ど|にち)ようび'
ENUM = re.compile(rf'^(?:{_NUM}+{_CNT}|{_DAY})$')

def is_enum_item(it):
    ch = it['questions'][0]['choices']
    return all(ENUM.match(strip_furi(c)) for c in ch)
CJSON = os.path.join(ROOT, 'content', 'problems', 'choukai')
NAME = {'kadai': '課題理解', 'point': 'ポイント理解', 'gaiyou': '概要理解', 'hatsuwa': '発話表現', 'sokuji': '即時応答'}
CONTENT = {'kadai', 'point', 'gaiyou'}   # テキスト4択＝語彙マッチ・設問テンプレが有効


def band_for(cat, lv):
    return SB.BAND[lv] if cat == 'sokuji' else BAND.get(f'{cat}_{lv}', (0, 9999))

def toks(t):
    return set(TOK.findall(strip_furi(t)))

def stem_norm(s):
    return re.sub(r'\s', '', strip_furi(s))


def analyze(cat, lv, items):
    n = len(items)
    audioch = bool(items[0].get('audioChoices'))
    nch = statistics.mode([len(it['questions'][0]['choices']) for it in items])
    scr_m = [body_mora(cat, it['script']) for it in items]
    lo, hi = band_for(cat, lv)
    short = sum(1 for m in scr_m if m < lo); long = sum(1 for m in scr_m if m > hi)
    hitLong = 0; POS = Counter(); kak = rusu = 0; chspread = []
    lexWin = 0; ovC = []; ovD = []; stems = Counter()
    reqhit = sephit = 0; KI = Counter(); AX = Counter()
    cores = []; chsets = []
    for it in items:
        q = it['questions'][0]; ch = q['choices']; ai = q['answerIndex']
        cm = [body_mora('sokuji', c) for c in ch]
        chspread.append(max(cm) - min(cm))
        if max(range(len(cm)), key=lambda k: cm[k]) == ai: hitLong += 1
        POS[ai + 1] += 1
        blob = strip_furi(it['script']) + ''.join(strip_furi(c) for c in ch)
        if KAKARI.search(blob): kak += 1
        if any(w in blob for w in ACCENT): rusu += 1
        cores.append(strip_furi(body_text(cat, it['script'])))   # 会話本文のみ（テンプレ導入ナレを除外）
        if not is_enum_item(it):   # 列挙型(数量/曜日)の答えは選択肢セット重複の対象外＝偽陽性防止
            chsets.append(" ".join(sorted(strip_furi(c) for c in ch)))
        if cat in CONTENT:
            body = toks(body_text(cat, it['script']))
            ov = [len(toks(c) & body) for c in ch]
            if ov[ai] > max(ov[k] for k in range(len(ov)) if k != ai): lexWin += 1
            ovC.append(ov[ai]); ovD.append(statistics.mean([ov[k] for k in range(len(ov)) if k != ai]))
            stems[stem_norm(q.get('q', ''))] += 1
        if cat == 'hatsuwa':
            forms = [_form(c) for c in ch]
            ri = [k for k, ff in enumerate(forms) if ff in REQ]
            if len(ri) == 1 and ri[0] == ai: reqhit += 1
            if forms[ai] not in [forms[k] for k in range(len(forms)) if k != ai]: sephit += 1
            KI[classify_function(it)] += 1; AX[classify_axis(it)] += 1
    scrdup = sum(1 for i, c in enumerate(cores) if (cores[:i] + cores[i+1:]) and nearest(c, cores[:i] + cores[i+1:])[0] >= DUP_SCRIPT)
    chdup = sum(1 for i, c in enumerate(chsets) if (chsets[:i] + chsets[i+1:]) and nearest(c, chsets[:i] + chsets[i+1:])[0] >= DUP_CHSET)
    return {
        '大問': NAME[cat], 'レベル': lv, '問題数': n, '選択肢数': nch,
        '選択肢方式': '音声焼込み(位置固定)' if audioch else '表示時ランダム',
        '台本モーラ min/中/max': f"{min(scr_m)}/{int(statistics.median(scr_m))}/{max(scr_m)}",
        '設計帯': f"{lo}-{hi}", '帯外 短/長': f"{short}/{long}",
        '最長を選ぶ的中%': round(hitLong / n * 100), '基準%': round(100 / nch),
        '語彙マッチ的中%': round(lexWin / n * 100) if cat in CONTENT else '',
        '正解−誤答 本文一致差': round(statistics.mean(ovC) - statistics.mean(ovD), 1) if cat in CONTENT else '',
        '設問テンプレ最大%': round(max(stems.values()) / n * 100) if cat in CONTENT else '',
        '台本重複(≥.60)': scrdup, '選択肢セット重複(≥.70)': chdup,
        '正解位置(正本)': "/".join(str(POS.get(k, 0)) for k in range(1, nch + 1)),
        '位置最大%(有効時)': round(max(POS.values()) / n * 100) if audioch else '—',
        '係': kak, '留守': rusu,
        '依頼形%': round(reqhit / n * 100) if cat == 'hatsuwa' else '',
        '形分離%': round(sephit / n * 100) if cat == 'hatsuwa' else '',
        '機能max%': round(max(KI.values()) / n * 100) if cat == 'hatsuwa' else '',
        '弁別軸max%': round(max(AX.values()) / n * 100) if cat == 'hatsuwa' else '',
    }


def load_all():
    g = {}
    for f in sorted(glob.glob(os.path.join(CJSON, '*.json'))):
        for it in json.load(open(f, encoding='utf-8'))['items']:
            g.setdefault((it.get('subtype') or 'kadai', it['level']), []).append(it)
    return g

COLS = ['大問', 'レベル', '問題数', '選択肢数', '選択肢方式', '台本モーラ min/中/max', '設計帯', '帯外 短/長',
        '最長を選ぶ的中%', '基準%', '語彙マッチ的中%', '正解−誤答 本文一致差', '設問テンプレ最大%',
        '台本重複(≥.60)', '選択肢セット重複(≥.70)', '正解位置(正本)', '位置最大%(有効時)', '係', '留守',
        '依頼形%', '形分離%', '機能max%', '弁別軸max%']


def rows_all():
    g = load_all()
    rows = []
    for cat in ['kadai', 'point', 'gaiyou', 'hatsuwa', 'sokuji']:
        for lv in ['N3', 'N4', 'N5']:
            if (cat, lv) in g:
                rows.append(analyze(cat, lv, g[(cat, lv)]))
    return rows


def pr(r):
    line = (f"{r['大問']:6s}{r['レベル']} n={r['問題数']:3d} {r['選択肢数']}択 {r['選択肢方式']:16s} "
            f"mora{r['台本モーラ min/中/max']:>11s} 帯{r['設計帯']:>7s} 帯外{r['帯外 短/長']:>5s} "
            f"最長{r['最長を選ぶ的中%']:2d}%(基{r['基準%']}) ")
    if r['大問'] in ('課題理解', 'ポイント理解', '概要理解'):
        line += f"語彙マッチ{r['語彙マッチ的中%']:2d}% 正誤差{r['正解−誤答 本文一致差']:+.1f} 設問{r['設問テンプレ最大%']}% "
    line += f"台本重複{r['台本重複(≥.60)']} 選択肢重複{r['選択肢セット重複(≥.70)']} 位置{r['正解位置(正本)']}最大{r['位置最大%(有効時)']} 係{r['係']}留守{r['留守']}"
    if r['大問'] == '発話表現':
        line += f" 依頼形{r['依頼形%']}%形分離{r['形分離%']}%機能{r['機能max%']}%軸{r['弁別軸max%']}%"
    print(line)


def write_xlsx(rows):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    xlsx = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
    wb = load_workbook(xlsx)
    sh = '聴解攻略耐性分析'
    if sh in wb.sheetnames: del wb[sh]
    ws = wb.create_sheet(sh)
    notes = [
        '聴解 攻略耐性・モーラ・ワンパターン 分析（2026-08-16 自動集計｜daimon_solvability.py）',
        '本質3観点→ ①答えが本文/設問から予測できる ②誤答が弱い ③大問×レベル内でワンパターン/重複。',
        '語彙マッチ的中%=「本文と最も語が重なる選択肢」を選ぶと正解になる割合(基準=1/選択肢数)。高い=正解だけ本文語→聞かず語マッチで解ける(①②)。良い設計は正解を言い換え本文語は誤答へ罠として置く。',
        '正解−誤答 本文一致差=正解の本文一致語数−誤答平均。+大=正解だけ本文語(①②)。最長を選ぶ的中=正解が最長選択肢の割合(②)。',
        '選択肢方式「表示時ランダム」(課題/ポイント/概要=テキスト4択)はアプリが並べ替え→正解位置は無関係(正本①固定でOK)。「音声焼込み」(発話/即時)は位置固定→位置偏りが有効。',
        '台本重複/選択肢セット重複=同一大問×レベル内の近似(③)。設問テンプレ最大=同一設問文の最大シェア(③)。設計帯:課題/ポイント/概要/発話=公式中央値±20%、即時=公式実測再センタリング。既存問の帯外は帯導入前の据え置き。',
    ]
    for tx in notes: ws.append([tx])
    ws.append(COLS)
    hr = len(notes) + 1
    for c in ws[hr]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDEBF7'); c.alignment = Alignment(wrap_text=True, vertical='center')
    for r in rows: ws.append([r.get(c, '') for c in COLS])
    widths = [12, 6, 6, 6, 16, 15, 8, 8, 12, 6, 12, 14, 12, 12, 14, 12, 12, 5, 5, 8, 8, 8, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f'A{hr + 1}'
    wb.save(xlsx)
    print(f"\n書込み: シート「{sh}」 {len(rows)}行 → {xlsx}")


def gate_draft(path, cat, lv):
    """作問ドラフト(items配列 or {'items':[...]})を採点＝将来作問QA。"""
    d = json.load(open(path, encoding='utf-8'))
    items = d.get('items', d) if isinstance(d, dict) else d
    r = analyze(cat, lv, items)
    pr(r)
    base = r['基準%']
    warns = []
    if cat in CONTENT:
        if r['語彙マッチ的中%'] > base + 10: warns.append(f"語彙マッチ{r['語彙マッチ的中%']}%>基準+10(聞かず語マッチで解ける恐れ)")
        if r['正解−誤答 本文一致差'] > 1.0: warns.append(f"正解−誤答差{r['正解−誤答 本文一致差']:+}>+1.0(正解だけ本文語)")
        if r['設問テンプレ最大%'] > 30: warns.append(f"設問テンプレ{r['設問テンプレ最大%']}%>30(ワンパターン)")
    if r['最長を選ぶ的中%'] > base + 15: warns.append(f"最長{r['最長を選ぶ的中%']}%>基準+15(長さで解ける)")
    if r['台本重複(≥.60)']: warns.append(f"台本重複{r['台本重複(≥.60)']}件")
    if r['選択肢セット重複(≥.70)']: warns.append(f"選択肢セット重複{r['選択肢セット重複(≥.70)']}件")
    print("  " + ("⚠ " + " / ".join(warns) if warns else "✅攻略耐性・ワンパターンとも良好"))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', action='store_true')
    ap.add_argument('--draft', nargs=3, metavar=('FILE', 'CAT', 'LV'))
    a = ap.parse_args()
    if a.draft:
        gate_draft(a.draft[0], a.draft[1], a.draft[2]); return
    rows = rows_all()
    for r in rows: pr(r)
    if a.xlsx: write_xlsx(rows)


if __name__ == '__main__':
    main()
