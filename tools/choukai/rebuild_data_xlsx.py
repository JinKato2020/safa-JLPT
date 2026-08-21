# -*- coding: utf-8 -*-
r"""聴解データ_{N5,N4,N3}.xlsx（ブラインド検品用ワークブック）を content JSON から再生成＝恒久ツール。

【なぜ】`問題/聴解/聴解データ_{lv}.xlsx`＝台本＋選択肢だけを見て一意性を人手確認する検品簿
（§即時応答 層3／§発話 ブラインド確認）。生成器が無く手更新で陳腐化していた→content正本から
全シートを機械再生成して「最新の問題文・選択肢・タグ」に揃える。既存の列レイアウト・見出し行の
書式は保持し、データ行(2行目以降)の値だけを id 突き合わせで上書き（行数はcontentに合わせて増減）。

シート別 列マップ（既存ファイルの見出しに一致）:
  課題理解  : 本文ID/シナリオ/観点(develop)/スクリプト/設問/選択肢1-4/正解
  ポイント理解: 本文ID/シナリオ/観点(kanten)/スクリプト/設問/選択肢1-4/正解
  概要理解  : 本文ID/シナリオ/ジャンル/設問型/スクリプト/設問/選択肢1-4/正解   （N3のみ）
  発話表現  : 本文ID/機能/場面/軸/状況スクリプト/選択肢1-3/正解
  即時応答  : 本文ID/機能/応答型/投げかけスクリプト/選択肢1-3/正解
正解列＝「選択肢{answerIndex+1}」（課題/ポイント/概要はanswerIndex=0固定＝選択肢1／発話・即時は焼込み位置）。

使い方: python tools/choukai/rebuild_data_xlsx.py [N5 N4 N3]   # 省略時=3レベル全部
※ 対象xlsxを Excel で開いていると PermissionError→閉じてから再実行（LOCKED表示）。
"""
import os, sys, io, json, glob
import openpyxl
from openpyxl.styles import Alignment

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CJSON = os.path.join(BASE, 'content', 'problems', 'choukai')
XDIR = os.path.join(BASE, '問題', '聴解')
LEVELS = [a for a in sys.argv[1:] if a in ('N5', 'N4', 'N3')] or ['N5', 'N4', 'N3']

# シート名 -> (subtype, [列定義])。列定義＝(見出し, 値の取り出し方)。'ch{n}'=選択肢n(1起点)。
SHEETS = {
    '課題理解': ('kadai', ['本文ID', 'シナリオ', '観点(develop)', 'スクリプト', '設問', '選択肢1', '選択肢2', '選択肢3', '選択肢4', '正解']),
    'ポイント理解': ('point', ['本文ID', 'シナリオ', '観点(kanten)', 'スクリプト', '設問', '選択肢1', '選択肢2', '選択肢3', '選択肢4', '正解']),
    '概要理解': ('gaiyou', ['本文ID', 'シナリオ', 'ジャンル', '設問型', 'スクリプト', '設問', '選択肢1', '選択肢2', '選択肢3', '選択肢4', '正解']),
    '発話表現': ('hatsuwa', ['本文ID', '機能', '場面', '軸', '状況スクリプト', '選択肢1', '選択肢2', '選択肢3', '正解']),
    '即時応答': ('sokuji', ['本文ID', '応答型は下で', '', '', '', '', '', '']),  # 即時は特別扱い（下）
}
# 即時応答の列は上のプレースホルダでなく実列で定義
SOKUJI_COLS = ['本文ID', '機能', '応答型', '投げかけスクリプト', '選択肢1', '選択肢2', '選択肢3', '正解']


def load_items():
    by_sub = {}
    for f in glob.glob(os.path.join(CJSON, '*.json')):
        d = json.load(io.open(f, encoding='utf-8'))
        for it in d.get('items', []):
            by_sub.setdefault((it.get('subtype') or 'kadai', it['level']), []).append(it)
    return by_sub


def cell_val(col, it):
    q = it['questions'][0]
    ch = q.get('choices', [])
    ai = q.get('answerIndex', 0)
    if col == '本文ID':
        return it['id']
    if col == 'シナリオ':
        return it.get('scenario', '')
    if col == '観点(develop)':
        return it.get('develop', '')
    if col == '観点(kanten)':
        return it.get('kanten', '')
    if col == 'ジャンル':
        return it.get('genre', '')
    if col == '設問型':
        return it.get('q_type', '')
    if col == '機能':
        return it.get('function', '')
    if col == '場面':
        return it.get('scene', '')
    if col == '軸':
        return it.get('axis', '')
    if col == '応答型':
        return it.get('answer_type', '')
    if col in ('スクリプト', '状況スクリプト', '投げかけスクリプト'):
        return it.get('script', '')
    if col == '設問':
        return q.get('q', '')
    if col.startswith('選択肢'):
        i = int(col[-1]) - 1
        return ch[i] if i < len(ch) else ''
    if col == '正解':
        return f'選択肢{ai + 1}'
    return ''


wrap = Alignment(wrap_text=True, vertical='top')
summary = []
for lv in LEVELS:
    p = os.path.join(XDIR, f'聴解_データ_{lv}.xlsx') if not os.path.exists(os.path.join(XDIR, f'聴解データ_{lv}.xlsx')) else os.path.join(XDIR, f'聴解データ_{lv}.xlsx')
    if not os.path.exists(p):
        summary.append(f'{lv}: ファイル無し {p}'); continue
    by_sub = load_items()
    wb = openpyxl.load_workbook(p)
    total = 0
    for sn in wb.sheetnames:
        if sn not in SHEETS:
            continue
        sub = SHEETS[sn][0]
        cols = SOKUJI_COLS if sn == '即時応答' else SHEETS[sn][1]
        ws = wb[sn]
        # 見出しは既存を尊重（実列見出しでマップ）。実際の見出し順を読む。
        hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        cols = [h for h in hdr if h]  # 実ファイルの見出し順に合わせる
        items = sorted(by_sub.get((sub, lv), []), key=lambda x: x['id'])
        # 既存データ行を一旦クリア（値のみ・2行目以降）
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).value = None
        # 書き直し
        for ri, it in enumerate(items, start=2):
            for ci, col in enumerate(cols, start=1):
                cell = ws.cell(ri, ci)
                cell.value = cell_val(col, it)
                cell.alignment = wrap
        total += len(items)
    try:
        wb.save(p); summary.append(f'{lv}: OK 更新 {total} 行 -> {p}')
    except PermissionError:
        summary.append(f'{lv}: LOCKED(Excelで開いています)→閉じて再実行')
    wb.close()
print('\n'.join(summary))
