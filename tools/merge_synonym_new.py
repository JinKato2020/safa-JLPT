# 未マージの言い換え類義EXCEL問題を content へマージする(追加のみ・既存は不変)。
#   N5(+100)/N4(+181)= 級内確定ファイル(現contentを含む上位集合)から vocabId未収録の行だけ追加。
#   N3(+554)= 新規問題(295)+増作2(259)。explain: 新規=Excel(半角→全角ルビ変換), 増作2=フリガナworkflow出力(fout_*.json)。
#   スキーマは級ごとに既存itemへ厳密一致(N5=stem+pattern・i18nなし / N4=stem・i18nなし / N3=i18n.ja.explain・stem/patternなし)。
import json, os, glob, re
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OD = r'C:\Users\jwpsa\AppData\Local\Temp\claude\c--Users-jwpsa-Documents-desktop-claude-JLPT---\dff2926b-efae-47c7-a4fd-57447b7e8983\scratchpad\syn_n3'
def sheet(path, name):
    wb = load_workbook(os.path.join(ROOT, path), read_only=True); ws = wb[name]
    rr = list(ws.iter_rows(values_only=True)); hdr = list(rr[0])
    out = [{hdr[i]: r[i] for i in range(len(hdr))} for r in rr[1:] if any(c not in (None, '') for c in r)]
    wb.close(); return out
def load_content(lv):
    p = os.path.join(ROOT, f'content/problems/moji_goi/synonym_{lv}.json')
    return p, json.load(open(p, encoding='utf-8'))
def to_full(s):  # 半角()→全角（）(ふりがな括弧の統一)
    return (s or '').replace('(', '（').replace(')', '）')
def idmax(items, pat):
    n = [int(re.search(r'(\d+)$', it['id']).group(1)) for it in items if re.search(pat, it['id'])]
    return max(n) if n else 0
def choices(row, keys):
    return [row[k] for k in keys if row.get(k) not in (None, '')]

report = {}

# ---------- N5 ----------
p5, d5 = load_content('N5'); cont5 = {it['vocabId'] for it in d5['items']}
k = idmax(d5['items'], r'N5-V-I'); added = 0
for r in sheet('言い換え類義_N5_級内確定_231問.xlsx', 'N5 級内確定'):
    if r['vocabId'] in cont5: continue
    cont5.add(r['vocabId']); k += 1
    d5['items'].append({'id': f'N5-V-I-{k:04d}', 'vocabId': r['vocabId'], 'sentence': r['sentence'],
        'underline': r['word'], 'word': r['word'], 'answer': r['answer'],
        'choices': choices(r, ['choice1','choice2','choice3']), 'verified': True,
        'stem': r['sentence'], 'pattern': r.get('pattern') or 'noun'}); added += 1
report['N5'] = {'added': added, 'total': len(d5['items'])}

# ---------- N4 ----------
p4, d4 = load_content('N4'); cont4 = {it['vocabId'] for it in d4['items']}
k = idmax(d4['items'], r'N4-V-I'); added = 0
for r in sheet('言い換え類義_N4_級内確定_406問.xlsx', 'N4 級内確定'):
    if r['vocabId'] in cont4: continue
    cont4.add(r['vocabId']); k += 1
    d4['items'].append({'id': f'N4-V-I-{k:04d}', 'vocabId': r['vocabId'], 'sentence': r['sentence'],
        'underline': r['word'], 'word': r['word'], 'answer': r['answer'],
        'choices': choices(r, ['choice1','choice2','choice3','choice4','choice5','choice6']),
        'verified': True, 'stem': r['sentence']}); added += 1
report['N4'] = {'added': added, 'total': len(d4['items'])}

# ---------- N3 ----------
p3, d3 = load_content('N3'); cont3 = {it['vocabId'] for it in d3['items']}
k = idmax(d3['items'], r'N3-V-I')
# 増作2 explain のフリガナ出力
fexp = {}
for f in glob.glob(os.path.join(OD, 'fout_*.json')):
    for e in json.load(open(f, encoding='utf-8')): fexp[e['id']] = e.get('explain_ruby') or e.get('explain')
shin = sheet('言い換え類義_新規問題_N5-100_N4-159_N3-300.xlsx', 'N3 言い換え類義')
zou  = sheet('言い換え類義_増作2_N5-36_N4-44_N3-272.xlsx', 'N3 言い換え類義')
added = 0; miss_fur = []
for src, r in [('shin', x) for x in shin] + [('zou', x) for x in zou]:
    if r['vocabId'] in cont3: continue
    cont3.add(r['vocabId']); k += 1
    if src == 'shin':
        explain = to_full(r.get('explain_ja'))
    else:
        ex = fexp.get(r['id'])
        if not ex: miss_fur.append(r['id']); ex = to_full(r.get('explain_ja'))
        explain = ex
    d3['items'].append({'id': f'N3-V-I-{k:04d}', 'vocabId': r['vocabId'],
        'i18n': {'ja': {'explain': explain}}, 'sentence': r['sentence'], 'underline': r['underline'],
        'word': r['word'], 'answer': r['answer'],
        'choices': choices(r, ['choice1','choice2','choice3']), 'verified': True}); added += 1
report['N3'] = {'added': added, 'total': len(d3['items']), 'missing_furigana': len(miss_fur)}

for p, d in [(p5, d5), (p4, d4), (p3, d3)]:
    json.dump(d, open(p, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

print('MERGE REPORT')
for lv in ('N5','N4','N3'): print(' ', lv, report[lv])
