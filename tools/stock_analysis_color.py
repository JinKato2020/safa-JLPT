# -*- coding: utf-8 -*-
"""在庫・模試ストックまとめ.xlsx の分析4シートを更新＋信号色(緑=問題なし/黄=気になる/赤=危険/灰=対象外)。
   ・読解 品質パラメータ … 情報検索行を60問化の最新値へ更新(joho_figure_check相当)＋適用後/情報検索の品質セルを着色
   ・聴解攻略耐性分析     … 帯外/最長的中(基準比)/台本重複/選択肢セット重複を着色
   ・聴解 骨組みパラメータ分布 … 判定・最大%を着色
   ・単語×大問カバー率   … カバー率セルを着色＋「点検メモ」列に何が問題かを日本語で記載
   使い方: python tools/stock_analysis_color.py   （前提: 対象xlsxを閉じておくこと）
"""
import os, re, json, subprocess, sys, statistics as st, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import joho_solvability as js   # 走査性S/C・多様性の判定ロジックを単一ソースで再利用
sys.setrecursionlimit(1000000)   # openpyxl delete_rows がスタイル複製で深く再帰する対策
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
DK = os.path.join(ROOT, 'content', 'problems', 'dokkai')

GREEN='CDE8D4'; YELLOW='FCE7C0'; RED='F6C9C4'; GREY='E8E8E8'
def fill(c): return PatternFill('solid', fgColor=c)
LEG = '凡例：緑=問題なし ／ 黄=気になる ／ 赤=危険 ／ 灰=対象外・履歴'
def add_legend(ws):
    # 既存の凡例行を消してから1本だけ付け直す(再実行で重複しない)
    for mr in list(ws.merged_cells.ranges): ws.unmerge_cells(str(mr))  # 結合セルは delete_rows を壊す
    for r in range(ws.max_row,0,-1):
        if str(ws.cell(r,1).value or '').startswith('凡例：'): ws.delete_rows(r,1)
    ws.cell(ws.max_row+2,1, LEG).font = Font(italic=True, color='7A756C')

def pct(v):
    m = re.search(r'(-?\d+(?:\.\d+)?)\s*%', str(v)) if v is not None else None
    return float(m.group(1)) if m else None
def num(v):
    try: return float(str(v));
    except:
        m=re.search(r'-?\d+(?:\.\d+)?', str(v) if v is not None else ''); return float(m.group()) if m else None
def ab(v):  # "19/0" -> (19,0)
    m=re.findall(r'\d+', str(v) if v is not None else '')
    return (int(m[0]),int(m[1])) if len(m)>=2 else (None,None)

# ── 情報検索の最新実効字数(figure込み・ルビ除去)を計算 ──
def strip_ruby(s): return re.sub(r'\s','',re.sub(r'（[^）]*）','',s or ''))
def collect(v):
    if isinstance(v,str): return v
    if isinstance(v,list): return ''.join(collect(x) for x in v)
    if isinstance(v,dict): return ''.join(collect(x) for x in v.values())
    return ''
BAND={'N5':(200,375),'N4':(340,460),'N3':(510,690)}  # N4/N3=目標±15%(2026-08-21)
TGT ={'N5':'250[200-375]','N4':'400[340-460]','N3':'600[510-690]'}
# figure比率/図版依存%/本文自足% は正本ツール joho_figure_check.py の出力を採用（指標定義を統一）。
fc = subprocess.run([sys.executable,'-X','utf8', os.path.join(ROOT,'tools','joho_figure_check.py')],
                    capture_output=True, text=True, encoding='utf-8').stdout
fcm={}  # lv -> (figratio, izon, jisoku)  ※行例: N5 60 60 271(204/313) 250[150-375] 0/0 73% 31% 68% 0
for ln in fc.splitlines():
    m=re.match(r'^\s*(N[345])\b.*?(\d+)%\s+(\d+)%\s+(\d+)%\s+\d+\s*$', ln)
    if m: fcm[m.group(1)]=(int(m.group(2)),int(m.group(3)),int(m.group(4)))
joho={}
for lv in ('N5','N4','N3'):
    its=json.load(open(os.path.join(DK,f'joho_{lv}.json'),encoding='utf-8'))['items']
    eff=[len(strip_ruby(collect(it.get('body'))+collect(it.get('figure')))) for it in its]
    lo,hi=BAND[lv]
    short=sum(1 for c in eff if c<lo); long=sum(1 for c in eff if c>hi)
    figratio,izon,jisoku = fcm.get(lv,(None,None,None))
    # 走査性・多様性(番人 joho_solvability と同基準)
    sc=collections.Counter(it.get('skeleton',{}).get('scene') for it in its)
    kinds=len([k for k in sc if k not in ('その他',None)]); top,topc=sc.most_common(1)[0]
    div=f"{kinds}種/最頻{round(topc/len(its)*100)}%"
    hard=lv in ('N4','N3')   # 走査S/Cのハード対象(N5は易しく=対象外)
    sbad=sum(1 for it in its if not js.sources_ok(it)[0])
    cbad=sum(1 for it in its if not js.choices_ok(it)[0])
    scanS='OK' if sbad==0 else f'NG{sbad}'
    scanC='OK' if cbad==0 else f'NG{cbad}'
    if not hard: scanS='対象外'; scanC='対象外'
    joho[lv]=dict(n=len(its), med=int(st.median(eff)), mn=min(eff), mx=max(eff),
                  tgt=TGT[lv], out=f'{short}/{long}', figratio=figratio, izon=izon, jisoku=jisoku,
                  div=div, scanS=scanS, scanC=scanC)

wb=load_workbook(XLSX)

# ══ Sheet: 読解 品質パラメータ ══
ws=wb['読解 品質パラメータ']
for mr in list(ws.merged_cells.ranges): ws.unmerge_cells(str(mr))  # delete_rows前に結合解除
# (1) 古い「現行」行を削除(ユーザー指示2026-08-21)。下から消して行番号ズレを防ぐ。
for r in range(ws.max_row,0,-1):
    if str(ws.cell(r,3).value or '').startswith('現行'): ws.delete_rows(r,1)
# (2) 残った「適用後(+N)」行＝現在の中身 → 区分を「最新」に付け替え
for r in range(1,ws.max_row+1):
    if str(ws.cell(r,3).value or '').startswith('適用後'): ws.cell(r,3,'最新')
# (2.5) 内容理解の各指標を現在の中身から再計算して最新行へ反映(dokkai_solvability + 指示語)
_dk = subprocess.run([sys.executable,'-X','utf8', os.path.join(ROOT,'tools','dokkai_solvability.py')],
                     capture_output=True, text=True, encoding='utf-8').stdout.splitlines()
dkm={}
for i,ln in enumerate(_dk):
    m=re.match(r'^(内容理解[短中長])\s+(N[345])\s+(\d+)\s+(\d+)\(\s*(\d+)/\s*(\d+)\)\s+\S+\s+(\d+)/\s*(\d+)\s+(\d+)%\s+(\d+)%\s+(\d+)%', ln)
    if m:
        qt=_dk[i+1].split('設問型:',1)[1].strip() if i+1<len(_dk) and '設問型' in _dk[i+1] else ''
        dkm[(m.group(1),m.group(2))]=dict(setsu=int(m.group(3)),ji=f"{m.group(4)}({m.group(5)}/{m.group(6)})",
            obi=f"{m.group(7)}/{m.group(8)}",cho=f"{m.group(9)}%",maru=f"{m.group(10)}%",goi=f"{m.group(11)}%",qt=qt)
FILEMAP={'内容理解短':'naiyou_tan','内容理解中':'naiyou_chu','内容理解長':'choubun'}
def shiji_of(daimon,lv):
    fp=os.path.join(DK,f"{FILEMAP[daimon]}_{lv}.json")
    if not os.path.exists(fp): return (0,0,0)
    its=json.load(open(fp,encoding='utf-8'))['items']
    n=len(its); sj=sum(1 for it in its if re.search(r'指(す|し|して|している)', re.sub(r'（[^）]*）','',it['questions'][0].get('q',''))))
    return (n, sj, round(sj/n*100) if n else 0)
for r in range(1,ws.max_row+1):
    da=str(ws.cell(r,1).value or ''); lv=str(ws.cell(r,2).value or '')
    if da in FILEMAP and str(ws.cell(r,3).value or '')=='最新' and (da,lv) in dkm:
        d=dkm[(da,lv)]; n,sj,rate=shiji_of(da,lv)
        ws.cell(r,4,n); ws.cell(r,5,d['setsu']); ws.cell(r,6,f"{d['setsu']/n:.1f}({d['setsu']//n})" if n else '—')
        ws.cell(r,7,d['ji']); ws.cell(r,9,d['obi']); ws.cell(r,10,d['cho']); ws.cell(r,11,d['maru']); ws.cell(r,12,d['goi'])
        ws.cell(r,13,f"{sj}/{n}"); ws.cell(r,14,f"{rate}%"); ws.cell(r,15,d['qt'])
# (3) 見出し注記(現行/適用後の説明)を最新版に更新
for r in range(1,ws.max_row+1):
    if str(ws.cell(r,1).value or '').startswith('①内容理解'):
        ws.cell(r,1,'①内容理解（短/中/長）　本体JSONの現在値（旧「現行」比較行は削除済み）')
# (4) 情報検索データ行を動的特定(col A=級 かつ col B=数値)＋最新値に更新
#     列: A級 B掲示 C設問 D実効字数 E目標 F帯外 G figure比率 H図版依存% I本文自足%
info_rows={}
for r in range(1,ws.max_row+1):
    a=str(ws.cell(r,1).value or '').strip()
    if a in ('N5','N4','N3') and isinstance(ws.cell(r,2).value,(int,float)): info_rows[a]=r
for lv,r in info_rows.items():
    d=joho[lv]
    ws.cell(r,2,d['n']); ws.cell(r,3,d['n'])
    ws.cell(r,4,f"{d['med']}({d['mn']}/{d['mx']})")
    ws.cell(r,5,d['tgt']); ws.cell(r,6,d['out'])
    if d['figratio'] is not None:
        ws.cell(r,7,f"{d['figratio']}%"); ws.cell(r,8,f"{d['izon']}%"); ws.cell(r,9,f"{d['jisoku']}%")
    # 走査性・多様性(2026-08-21 追加・番人 johoSolvability と同基準)
    ws.cell(r,10,d['div']); ws.cell(r,11,d['scanS']); ws.cell(r,12,d['scanC'])
# 情報検索の見出し行(級/掲示…)に走査性・多様性の列見出しを追加
for r in range(1,ws.max_row+1):
    if str(ws.cell(r,1).value or '').strip()=='級' and str(ws.cell(r,2).value or '').strip()=='掲示':
        ws.cell(r,10,'多様性(種類/最頻)'); ws.cell(r,11,'走査S 情報源≥2'); ws.cell(r,12,'走査C 誘惑肢(選ぶ)')
        break
# (5) 着色: 内容理解「最新」行の品質セル  I(9)帯外 J(10)最長% K(11)丸写% L(12)語彙%
for r in range(1,ws.max_row+1):
    if not str(ws.cell(r,1).value or '').startswith('内容理解'): continue
    if str(ws.cell(r,3).value or '')!='最新': continue
    a,b=ab(ws.cell(r,9).value); tot=(a or 0)+(b or 0)
    ws.cell(r,9).fill=fill(GREEN if tot==0 else YELLOW if tot<=3 else RED)
    for c,(g,y) in {10:(25,35),12:(35,45)}.items():
        p=pct(ws.cell(r,c).value)
        if p is not None: ws.cell(r,c).fill=fill(GREEN if p<=g else YELLOW if p<=y else RED)
    p=pct(ws.cell(r,11).value)
    if p is not None: ws.cell(r,11).fill=fill(GREEN if p<=30 else YELLOW if p<=45 else RED)
# (6) 着色: 情報検索行  F帯外 / D実効字数(帯内=緑) / I本文自足%(高い=要点検)
for lv,r in info_rows.items():
    a,b=ab(ws.cell(r,6).value); tot=(a or 0)+(b or 0)
    ws.cell(r,6).fill=fill(GREEN if tot==0 else YELLOW if tot<=3 else RED)
    ws.cell(r,4).fill=fill(GREEN)
    p=pct(ws.cell(r,9).value)
    if p is not None: ws.cell(r,9).fill=fill(GREEN if p<=50 else YELLOW if p<=70 else RED)
    # 走査性・多様性の着色: OK/対象外=緑・NG=赤
    for c in (10,11,12):
        v=str(ws.cell(r,c).value or '')
        ws.cell(r,c).fill=fill(RED if v.startswith('NG') else GREEN)
add_legend(ws)

# ══ Sheet: 聴解攻略耐性分析 (header r7, data r8-20) ══
# 色基準の考え方(2026-08-21 recalibrate):
#  ・帯外(モーラ)は「帯導入前の既存問の据え置き」＝欠陥でない → 0のみ緑・それ以外は灰(対象外)。赤にしない。
#  ・本物の攻略余地＝最長選び的中(基準比)・台本/選択肢セット重複 のみ 黄/赤。
ws=wb['聴解攻略耐性分析']
for r in range(8,21):
    # H(8)帯外: 据え置き扱い(0=緑/他=灰)
    a,b=ab(ws.cell(r,8).value); tot=(a or 0)+(b or 0)
    ws.cell(r,8).fill = fill(GREEN if tot==0 else GREY)
    # I(9)最長を選ぶ的中% vs J(10)基準%  (差>15=赤/>5=黄)
    I=num(ws.cell(r,9).value); J=num(ws.cell(r,10).value)
    if I is not None and J is not None:
        d=I-J
        ws.cell(r,9).fill = fill(GREEN if d<=5 else YELLOW if d<=15 else RED)
    # N(14)台本重複  O(15)選択肢セット重複  (0=緑/<4=黄/≥4=赤)
    for col in (14,15):
        v=num(ws.cell(r,col).value)
        if v is not None: ws.cell(r,col).fill = fill(GREEN if v==0 else YELLOW if v<4 else RED)
add_legend(ws)

# ══ Sheet: 聴解 骨組みパラメータ分布 (header r5, data r6-13) ══
ws=wb['聴解 骨組みパラメータ分布']
for r in range(6,14):
    p=pct(ws.cell(r,8).value)                             # H 最大%
    if p is not None: ws.cell(r,8).fill = fill(GREEN if p<40 else YELLOW if p<=50 else RED)
    j=str(ws.cell(r,9).value or '')                       # I 判定
    if j: ws.cell(r,9).fill = fill(GREEN if '✅' in j else RED if '❌' in j else YELLOW)
add_legend(ws)

# ══ Sheet: 単語×大問カバー率 ══
ws=wb['単語×大問カバー率']
HDRS=[5,18,31]
for hr in HDRS:                                          # 「点検メモ」列(H=8)を追加
    ws.cell(hr,8,'点検メモ')
    ws.cell(hr,8).font=Font(bold=True,color='FFFFFF'); ws.cell(hr,8).fill=fill('2E5A88')
    ws.cell(hr,8).alignment=Alignment(horizontal='center',vertical='center')
ws.column_dimensions['H'].width=52
cur_shubetsu=''
for r in range(1,ws.max_row+1):
    A=ws.cell(r,1).value; B=ws.cell(r,2).value; F=ws.cell(r,6).value
    if A and str(A).strip().startswith(('語彙単語','漢字単語','文法単語')): cur_shubetsu=str(A).strip()
    if not B or r in HDRS: continue
    daimon=str(B).strip()
    bosuu=ws.cell(r,5).value; cov=ws.cell(r,4).value
    p=pct(F)
    if p is None:                                        # なし/— = 対象外
        if F not in (None,''):
            ws.cell(r,6).fill=fill(GREY)
            ws.cell(r,8,'用法は場面/例文単位で単語IDを持たない＝カバー率の対象外（欠落ではない）')
        continue
    if p>=80:
        ws.cell(r,6).fill=fill(GREEN)
    elif p>=60:
        ws.cell(r,6).fill=fill(YELLOW)
        ws.cell(r,8, f'カバー率{int(p)}%。{cur_shubetsu}の約{100-int(p)}%が「{daimon}」で一度も出題されない。余裕を持って補いたい。')
    else:
        ws.cell(r,6).fill=fill(RED)
        note=f'母数{bosuu}中{cov}問＝{int(p)}%。{cur_shubetsu}の約{100-int(p)}%が「{daimon}」で一度も出題されない。増産か別大問で補完を。'
        if p<=10: note+=' ※極端に低い＝データ欠落の疑いあり・要確認。'
        ws.cell(r,8, note)
    ws.cell(r,8).alignment=Alignment(wrap_text=True, vertical='top')
add_legend(ws)

wb.save(XLSX)
print('OK 更新+着色:', XLSX)
print('情報検索(最新):', {lv:f"{d['n']}問 med{d['med']}({d['mn']}/{d['mx']}) fig{d['figratio']}% 本文自足{d['jisoku']}%" for lv,d in joho.items()})
