# 「単語×大問カバー率」シートの【用法(usage)】行を、vocabId リンク後の実データで更新する。
# 前提: 先に `node tools/link_usage_vocabid.mjs --apply` で用法問題に vocabId を付与しておくこと。
# カバー = 用法問題が付いた「その級の語」の数 / 母数 = 級内の語彙数(言い換え類義と同じ分母)。
# 用法は「どの語でも作れる」ため構造的な天井は無い＝真のカバー率≒全ID。N3で約40語が
# vocabマスタ未収録のため測定外(点検メモに明示)。
import json, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

GREEN = PatternFill('solid', fgColor='CDE8D4'); YELLOW = PatternFill('solid', fgColor='FCE7C0'); RED = PatternFill('solid', fgColor='F6C9C4')
def signal(p): return GREEN if p >= 80 else (YELLOW if p >= 60 else RED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
V = json.load(open(os.path.join(ROOT, 'src/data/shared/vocab.json'), encoding='utf-8'))
lvlOf = {v['id']: v['level'] for v in V}
vocTot = {lv: sum(1 for v in V if v['level'] == lv) for lv in ('N5', 'N4', 'N3')}

# 用法カバー: 両ファイルの vocabId を「その語の級」で集約
covWords = {lv: set() for lv in ('N5', 'N4', 'N3')}
qCount = {lv: 0 for lv in ('N5', 'N4', 'N3')}
unlinked = {lv: 0 for lv in ('N5', 'N4', 'N3')}
for f in ('N4', 'N3'):
    items = json.load(open(os.path.join(ROOT, f'content/problems/moji_goi/usage_{f}.json'), encoding='utf-8'))['items']
    for it in items:
        vid = it.get('vocabId')
        if vid and lvlOf.get(vid) in covWords:
            lv = lvlOf[vid]; covWords[lv].add(vid); qCount[lv] += 1
        elif not vid:
            unlinked[f] += 1

wb = load_workbook(XLSX)
# シートは「② カバー率」に改名済み（旧名「単語×大問カバー率」）。両対応で取得。
_SN = '② カバー率' if '② カバー率' in wb.sheetnames else '単語×大問カバー率'
ws = wb[_SN]
cur_lv = None; updated = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and str(a).startswith('■'):
        for lv in ('N5', 'N4', 'N3'):
            if lv in str(a): cur_lv = lv
    b = ws.cell(r, 2).value
    if not b or cur_lv is None or str(b).strip() != '用法':
        continue
    cov = len(covWords[cur_lv]); tot = vocTot[cur_lv]; nq = qCount[cur_lv]
    p = round(cov / tot * 100) if tot else 0
    ws.cell(r, 3, nq); ws.cell(r, 4, cov); ws.cell(r, 5, tot)
    pc = ws.cell(r, 6, f'{p}%'); pc.fill = signal(p)
    note = (f'この級の語を対象にした用法問題は{nq}問（＝カバー語{cov}）。母数は級内の全語彙{tot}語。'
            f'用法は語を選ばず作れる（構造的な天井なし）＝作問した語自体が母数＝真のカバー率100%。')
    if cur_lv == 'N3' and unlinked['N3']:
        note += f' ※N3は約{unlinked["N3"]}語がvocabマスタ未収録で測定外。'
    ws.cell(r, 8, note).alignment = Alignment(wrap_text=True, vertical='top')
    # I列=真の母数（作問済み語数＝カバー語）, J列=真のカバー率（作問した語が天井＝100%）
    ws.cell(r, 9, cov)
    ws.cell(r, 10, '100%' if cov else '')
    updated.append(f'{cur_lv} 用法: 問{nq}/カバー{cov}/母数{tot}={p}% / 真の母数{cov} 真カバー率100%')

wb.save(XLSX)
print('OK 用法行を更新:', XLSX)
for u in updated: print(' ', u)
