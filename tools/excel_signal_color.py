# Excel の「状態セル」に信号色(緑=済/黄=一部/赤=未/灰=対象外)を値から自動で塗る。
# 目的: エクセル反映のたびに手で色を付け忘れる事故を仕組みで潰す(ユーザー厳命 2026-09-03)。
#   フック(PostToolUse Bash/PowerShell)が .xlsx 保存を検知して本スクリプトを自動起動する。手動起動も可。
#
# 【重要・安全設計】
#   ・塗るのは「手で状態を書くシート」だけ(下の SHEETS 許可リスト)。生成ツールが独自のしきい値で
#     色分けするシート(③品質・攻略耐性=攻略耐性%は低いほど良、②カバー率 等)には触れない。
#     許可リストに無いブックは何もしない(未知ブックを壊さない)。
#   ・既存の色は消さない(掃除しない)。生成ツールが付けた色を尊重する。該当する状態セルにだけ上塗り。
#   ・判定は「セルの文字列全体(前後空白除去)」。長い説明文・見出し・数値は一致せず塗られない。
#
# 使い方:
#   python tools/excel_signal_color.py                      # 既定=許可リストの全ブック
#   python tools/excel_signal_color.py path/to/a.xlsx ...   # 指定ブック(許可リストのシートだけ塗る)
#
#   ・パーセント: 先頭が数字+%(例 100%・62%・"100%（語義1182）") → 100↑=緑 / 1〜99=黄 / 0=赤
#   ・緑トークン: 完了 フル 元 対応 対応済 実装済 済 あり 有
#   ・黄トークン: 一部 途中 途上 進行中 部分 一部対応 backlog
#   ・赤トークン: 未着手 なし 未 未対応 無 ×
#   ・灰トークン(対象外/不要): — － - N/A 不要 訳なし 対象外
import sys, os, re, glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ブック(basename)→ 信号色を塗ってよいシート名。ここに無いブック/シートは一切触らない。
# 新しく「手で状態を書くシート」を作ったら 1 行足すだけ(それが仕組みの拡張点)。
SHEETS = {
    '在庫・模試ストックまとめ.xlsx': ['⑥ 翻訳状況'],
}

SIG = {
    'green': Font(color='FF006100'),
    'amber': Font(color='FF9C6500'),
    'red':   Font(color='FF9C0006'),
    'gray':  Font(color='FF808080'),
}
FILL = {
    'green': PatternFill('solid', fgColor='FFC6EFCE'),
    'amber': PatternFill('solid', fgColor='FFFFEB9C'),
    'red':   PatternFill('solid', fgColor='FFFFC7CE'),
    'gray':  PatternFill('solid', fgColor='FFF2F2F2'),
}

GREEN = {'完了', 'フル', '元', '対応', '対応済', '実装済', '済', 'あり', '有'}
AMBER = {'一部', '途中', '途上', '進行中', '部分', '一部対応', 'backlog'}
RED   = {'未着手', 'なし', '未', '未対応', '無', '×', 'x', 'X'}
GRAY  = {'—', '－', '-', 'ー', 'N/A', 'n/a', '不要', '訳なし', '対象外', '（不要）', '(不要)'}
PCT = re.compile(r'^(\d{1,3})\s*[%％]')


def classify(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    m = PCT.match(s)
    if m:
        n = int(m.group(1))
        return 'green' if n >= 100 else ('red' if n == 0 else 'amber')
    if s in GREEN:
        return 'green'
    if s in AMBER:
        return 'amber'
    if s in RED:
        return 'red'
    if s in GRAY:
        return 'gray'
    return None


def color_workbook(path):
    allow = SHEETS.get(os.path.basename(path))
    if not allow:
        print(f'[signal-color] 許可リスト外(スキップ): {os.path.basename(path)}')
        return None
    wb = load_workbook(path)
    painted = 0
    for name in allow:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                sig = classify(cell.value)
                if not sig:
                    continue  # 既存色は消さない。状態セルにだけ上塗り。
                cell.fill = FILL[sig]
                old = cell.font
                cell.font = Font(name=old.name, size=old.size, bold=old.bold,
                                 italic=old.italic, color=SIG[sig].color)
                painted += 1
    wb.save(path)
    print(f'[signal-color] {os.path.basename(path)} [{",".join(allow)}]: 塗={painted}')
    return painted


def targets(argv):
    args = [a for a in argv if a.lower().endswith('.xlsx')]
    if args:
        return args
    # 引数なし=許可リストのブックを memory/ から探す
    out = []
    for base in SHEETS:
        p = os.path.join(ROOT, 'memory', base)
        if os.path.exists(p):
            out.append(p)
    return out


def main():
    files = targets(sys.argv[1:])
    if not files:
        print('[signal-color] 対象xlsxなし'); return
    for f in files:
        if os.path.exists(f):
            color_workbook(f)
        else:
            print(f'[signal-color] 見つからず: {f}')


if __name__ == '__main__':
    main()
