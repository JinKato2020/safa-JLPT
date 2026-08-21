# N3「文章の文法」を確認用Excelへ出力（本文全文＋各空所の4択＋正解＋pointId）。
# 出典 = content/problems/bunpou/passage_grammar_N3.json（＝ビルドに乗っている正本・0001-0020）。
# 追加で out_n3b*.json（生成中の0021-0050）が有れば末尾に付ける（--include-drafts）。
import json, os, glob, sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'content', 'problems', 'bunpou')
GEN = os.path.join(ROOT, 'scratchpad', 'pg', 'gen')
OUTDIR = os.path.join(ROOT, '問題')
OUT = os.path.join(OUTDIR, 'N3_文章の文法_確認用_全50問.xlsx')

g = json.load(open(os.path.join(ROOT, 'src/data/shared/grammar.json'), encoding='utf-8'))
gmap = {x['id']: x for x in g}

wb = Workbook()
ws = wb.active
ws.title = 'N3 文章の文法'

hdr = Font(bold=True)
setfont = Font(bold=True, size=12)
green = PatternFill('solid', fgColor='C6EFCE')
grayfill = PatternFill('solid', fgColor='F2F2F2')
wrap = Alignment(wrap_text=True, vertical='top')
top = Alignment(vertical='top')

widths = [10, 6, 60, 22, 22, 22, 22, 8, 14, 22, 6]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

HEAD = ['セットID', '空所', '本文 / 選択肢', '選択肢1', '選択肢2', '選択肢3', '選択肢4',
        '正解', 'pointId', '文法点', '級']
ws.append(HEAD)
for c in ws[1]:
    c.font = hdr
    c.fill = grayfill

r = 2


def emit_set(s):
    global r
    sid = s['id']
    body = '\n'.join(p['body'] for p in s['passages'])
    # 本文行（セット見出し＋全文）
    ws.cell(r, 1, sid).font = setfont
    bc = ws.cell(r, 3, body)
    bc.alignment = wrap
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
    r += 1
    for q in sorted(s['questions'], key=lambda q: q['blankNo']):
        ai = q['answerIndex']
        ch = (q.get('choices') or []) + [''] * 4
        ws.cell(r, 2, q['blankNo']).alignment = top
        for j in range(4):
            cell = ws.cell(r, 4 + j, ch[j])
            cell.alignment = top
            if j == ai:
                cell.fill = green
                cell.font = Font(bold=True)
        ws.cell(r, 8, f'{ai + 1}（{ch[ai]}）').alignment = top
        pid = q.get('pointId', '')
        ws.cell(r, 9, pid).alignment = top
        gg = gmap.get(pid)
        ws.cell(r, 10, gg['point'] if gg else '').alignment = top
        ws.cell(r, 11, gg['level'] if gg else '').alignment = top
        r += 1


# 正本（live）
live = json.load(open(os.path.join(SRC, 'passage_grammar_N3.json'), encoding='utf-8'))
for s in live['items']:
    emit_set(s)
n_live = len(live['items'])

# 生成中ドラフト（任意）
n_draft = 0
if '--include-drafts' in sys.argv:
    seen = {s['id'] for s in live['items']}
    drafts = []
    for f in sorted(glob.glob(os.path.join(GEN, 'out_n3b*.json'))):
        d = json.load(open(f, encoding='utf-8'))
        drafts += d.get('sets', [])
    for s in sorted(drafts, key=lambda s: s['id']):
        if s['id'] in seen:
            continue
        seen.add(s['id'])
        emit_set(s)
        n_draft += 1

ws.freeze_panes = 'A2'
os.makedirs(OUTDIR, exist_ok=True)
wb.save(OUT)
print('WROTE', OUT)
print(f'live={n_live} draft={n_draft} rows={r - 1}')
