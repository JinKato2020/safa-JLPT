# サブエージェントの選定結果(usage_reduce_result_*.json)を用法N4に適用する。
#   1) 旧型5-7択 → [正解 + 残す3誤答] の4択に縮小(usage_N4.json)
#   2) サイドカー(usageDistractorTags.json)へ殺し方タグを追加・legacyAllowlistから除外
#   3) 一意性 borderline の誤答をオレンジ着色した確認用Excelを生成
# 検証NG(keepIdx<3 / repl重複 / 型が全同一など)は適用せず needs_manual に退避。
import json, io, os, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

SP = r'C:\Users\jwpsa\AppData\Local\Temp\claude\c--Users-jwpsa-Documents-desktop-claude-JLPT---\9572c53a-67a4-47fa-b5b2-80501cd6e84d\scratchpad'
JSON_N4 = 'content/problems/moji_goi/usage_N4.json'
SIDE = 'src/data/shared/usageDistractorTags.json'
XLSX = '用法N4_旧型4択化_確認用.xlsx'
ORANGE = PatternFill('solid', fgColor='FFC000')   # 一意性borderline
RED    = PatternFill('solid', fgColor='FF6B6B')   # 正解が怪しい(answerOk=false)
HEAD   = PatternFill('solid', fgColor='DDDDDD')

# --- 結果とバッチを収集 ---
batches, results = {}, {}
for i in range(1, 5):
    bp = os.path.join(SP, f'usage_reduce_batch_{i}.json')
    rp = os.path.join(SP, f'usage_reduce_result_{i}.json')
    for it in json.load(open(bp, encoding='utf-8')):
        batches[it['id']] = it
    if not os.path.exists(rp):
        print(f'!! 結果ファイル無し: {rp}'); sys.exit(1)
    for r in json.load(open(rp, encoding='utf-8')):
        results[r['id']] = r
print('batches', len(batches), 'results', len(results))

d = json.load(open(JSON_N4, encoding='utf-8'))
byid = {it['id']: it for it in d['items']}
side = json.load(open(SIDE, encoding='utf-8'))
mono_set = set(side.get('monoTypeAllow', []))
legacy_set = set(side.get('legacyAllowlist', []))

applied, needs_manual, rows = [], [], []
for iid, b in batches.items():
    r = results.get(iid)
    dist = b['distractors']
    reason = None
    if not r: reason = '結果欠落'
    else:
        ki = r.get('keepIdx', [])
        tg = r.get('tags', [])
        if len(ki) != 3 or len(set(ki)) != 3: reason = f'keepIdx不正{ki}'
        elif any((not isinstance(k, int)) or k < 0 or k >= len(dist) for k in ki): reason = f'keepIdx範囲外{ki}'
        elif len(tg) != len(ki): reason = 'タグ数不一致'
        else:
            repls = [t.get('repl', '') for t in tg]
            types = [t.get('type', '') for t in tg]
            if len(set(repls)) < 3: reason = f'repl重複(P1){repls}'
            elif len(set(types)) < 2 and not r.get('mono'): reason = f'型が全同一(P2){types}'
    if reason:
        needs_manual.append((iid, b['stem'], reason)); continue

    ki, tg = r['keepIdx'], r['tags']
    kept = [dist[k] for k in ki]
    it = byid[iid]
    it['choices'] = [it['answer']] + kept          # 正解を先頭に、残す3誤答
    side['tags'][iid] = [{'repl': t['repl'], 'type': t['type'],
                          'certainty': t.get('certainty', 'clear')} for t in tg]
    legacy_set.discard(iid)
    # monoは「本当に型が単一の時だけ」。エージェントが型混在でmono=trueにした分は解除(normal branchで通す)。
    if r.get('mono') and len(set(t['type'] for t in tg)) == 1:
        mono_set.add(iid)
    else:
        mono_set.discard(iid)
    applied.append(iid)
    rows.append({'id': iid, 'stem': b['stem'], 'answer': it['answer'],
                 'answerOk': r.get('answerOk', True), 'kept': kept, 'tags': tg,
                 'note': r.get('note', '')})

# --- 書き戻し ---
side['legacyAllowlist'] = sorted(legacy_set)
side['monoTypeAllow'] = sorted(mono_set)
json.dump(d, io.open(JSON_N4, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
json.dump(side, io.open(SIDE, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

# --- 確認用Excel(borderlineをオレンジ) ---
wb = openpyxl.Workbook(); ws = wb.active; ws.title = '用法N4 旧型4択化'
hdr = ['ID', '語', '○ 正解文', '誤答1', '型1', '誤答2', '型2', '誤答3', '型3', 'note']
ws.append(hdr)
for c in ws[1]: c.fill = HEAD; c.font = Font(bold=True)
ws.append(['凡例: オレンジ=一意性borderline(要確認) / 赤=正解が怪しい', '', '', '', '', '', '', '', '', ''])
border_cnt = 0
for row in rows:
    tg = row['tags']
    line = [row['id'], row['stem'], row['answer']]
    for t in tg:
        line += [t.get('_sent', ''), f"{t.get('type')}→{t.get('repl')}"]
    line.append(row['note'])
    ws.append(line)
    ri = ws.max_row
    # 誤答文セルにテキストを入れつつ borderline着色(誤答文は列4,6,8)
    for j, (sent, t) in enumerate(zip(row['kept'], tg)):
        col = 4 + j * 2
        ws.cell(ri, col).value = sent
        if t.get('certainty') == 'borderline':
            ws.cell(ri, col).fill = ORANGE; border_cnt += 1
    if not row['answerOk']:
        ws.cell(ri, 3).fill = RED
for col, w in zip('ABCDEFGHIJ', [13, 8, 40, 34, 12, 34, 12, 34, 12, 20]):
    ws.column_dimensions[col].width = w
for r_ in ws.iter_rows():
    for c in r_: c.alignment = Alignment(vertical='top', wrap_text=True)
wb.save(XLSX)

print(f'applied {len(applied)} / needs_manual {len(needs_manual)} / borderline誤答 {border_cnt}')
if needs_manual:
    print('--- needs_manual ---')
    for x in needs_manual: print('  ', x)
print('XLSX:', os.path.abspath(XLSX))
