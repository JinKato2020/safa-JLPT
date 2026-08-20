# -*- coding: utf-8 -*-
"""読解問題_N{3,4,5}.xlsx の「情報検索」シートだけを、現行 content の joho_N*.json で作り直す。
   内容理解(短/中/長)シートと、その区分(既存/新規)データには一切触れない（温存）。
   使い方: python tools/dokkai_joho_excel.py
   出力: 問題/読解/読解問題_N{3,4,5}.xlsx（情報検索シートのみ差し替え）
   ※ joho は 1掲示=1問。区分は今回作り直し分＝すべて「新規」。"""
import os, re, json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
LEVELS = ['N3', 'N4', 'N5']
SHEET = '情報検索'
HDR = ['区分', '本文ID', 'タイトル', '字数', '本文', '図版', '設問ID', '設問',
       '選択肢1', '選択肢2', '選択肢3', '選択肢4', '正解', '解説']

def eff_chars(body):
    return len(re.sub(r'\s', '', re.sub(r'（[^）]*）', '', body or '')))

def fig_text(fig):
    p = []
    for k in ('kind', 'header', 'intro'):
        if fig.get(k):
            p.append(str(fig[k]))
    for b in fig.get('blocks', []):
        t = b.get('type')
        if t == 'table':
            p.append('table')
            if b.get('title'):
                p.append(str(b['title']))
            tb = b.get('table', {})
            p += [str(c) for c in tb.get('columns', [])]
            for row in tb.get('rows', []):
                p += [str(c) for c in row]
        elif t in ('notice', 'card'):
            p.append(t)
            if b.get('title'):
                p.append(str(b['title']))
            p += [str(l) for l in b.get('lines', [])]
    for n in fig.get('notes', []):
        p.append(str(n))
    if fig.get('footer'):
        p.append(str(fig['footer']))
    return ' / '.join(p)

def main():
    for lv in LEVELS:
        xls = os.path.join(ROOT, '問題', '読解', f'読解問題_{lv}.xlsx')
        jsn = os.path.join(ROOT, 'content', 'problems', 'dokkai', f'joho_{lv}.json')
        data = json.load(open(jsn, encoding='utf-8'))
        wb = load_workbook(xls)
        if SHEET in wb.sheetnames:
            idx = wb.sheetnames.index(SHEET)
            wb.remove(wb[SHEET])
        else:
            idx = len(wb.sheetnames)
        ws = wb.create_sheet(SHEET, idx)
        # ヘッダー
        ws.append(HDR)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='DDDDDD')
            c.alignment = Alignment(vertical='center')
        # データ
        n = 0
        for it in data['items']:
            q = it['questions'][0]
            ch = q['choices']
            ws.append([
                '新規', it['id'], it.get('title', ''), eff_chars(it.get('body', '')),
                it.get('body', ''), fig_text(it.get('figure', {})),
                q['id'], q['q'], ch[0], ch[1], ch[2], ch[3],
                f"選択肢{q['answerIndex'] + 1}", q['i18n']['ja']['explain'],
            ])
            n += 1
        # 見やすさ: 折り返し
        widths = {'A': 6, 'B': 14, 'C': 22, 'D': 6, 'E': 48, 'F': 60, 'G': 16, 'H': 30,
                  'I': 20, 'J': 20, 'K': 20, 'L': 20, 'M': 8, 'N': 48}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical='top')
        wb.save(xls)
        print(f"{lv}: 情報検索シート再生成 {n}問 -> {xls}")

if __name__ == '__main__':
    main()
