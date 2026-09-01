# -*- coding: utf-8 -*-
"""問題を「一意性チェック用」Excelへ書き出す（レベル毎ファイル・大問毎シート）。

出力（セッション＝プロジェクト直下）:
  一意性チェック_模試_N5.xlsx / _N4 / _N3   … pool=mock の大問だけ
  一意性チェック_通常_N5.xlsx / _N4 / _N3   … pool=normal の大問

色付け（一意性リスク）:
  🔴 赤   = 別解の恐れ大（uniqRisk=high、または ambiguous=true）
  🟡 黄   = 要確認（uniqRisk=mid、または verified=false）
  無色    = 一意で問題なし

★ 色の正本＝各問の自己申告フィールド（生成エージェントが本文と同時に付ける）:
    "uniqRisk": "high" | "mid"   （無ければ無リスク）
    "uniqNote": "<なぜあやしいかの理由>"
  既存問題はこの欄が無いので、暫定で verified=false→黄 / ambiguous=true→赤 を写す。

使い方: python tools/quality_excel.py
"""
import io, os, re, glob, json, sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
SRC = os.path.join(ROOT, 'content', 'problems')
LEVELS = ['N5', 'N4', 'N3']

RED = 'F6C9C4'   # 別解の恐れ大
AMBER = 'FCE7C0'  # 要確認
HEAD = '2E5A88'
thin = Side(style='thin', color='C9D3DE')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

COLS = ['問題ID', '種別', '本文/読み物', '設問', '対象語', '選択肢', '正解', '一意性リスク', '理由', 'verified']
WIDTHS = [22, 12, 52, 40, 14, 46, 20, 12, 34, 10]

# 本文/読み物として拾うキー（優先順）。読解=body/passages・聴解=script（scenarioは短い場面ラベルなので後回し）・
# 文字語彙文法の単文=sentence/prompt/stem。※「本文がメイン」ゆえ運搬文を必ず出す（ユーザー厳命）。
BODY_KEYS = ['body', 'passages', 'passage', 'script', 'sentence', 'prompt', 'stem', 'scenario', 'text', 'reading', 'title']
# 設問（汎用の指示文）。本文に使う prompt/stem とは分けて question/q だけを拾う。
Q_KEYS = ['question', 'q']

# 大問キー→日本語の大問名（シート名に使う）
JP_NAME = {
    'kanji_read': '漢字読み', 'orthography': '表記', 'context': '文脈規定',
    'synonym': '言い換え類義', 'usage': '用法',
    'grammar_form': '文法形式の判断', 'order': '文の組み立て', 'passage_grammar': '文章の文法',
    'naiyou_tan': '内容理解（短文）', 'naiyou_chu': '内容理解（中文）',
    'choubun': '内容理解（長文）', 'joho': '情報検索',
    'kadai': '課題理解', 'point': 'ポイント理解', 'gaiyou': '概要理解',
    'hatsuwa': '発話表現', 'sokuji': '即時応答',
}
# 本番試験の出題順（言語知識[文字語彙]→[文法]→読解→聴解）
EXAM_ORDER = [
    'kanji_read', 'orthography', 'context', 'synonym', 'usage',
    'grammar_form', 'order', 'passage_grammar',
    'naiyou_tan', 'naiyou_chu', 'choubun', 'joho',
    'kadai', 'point', 'gaiyou', 'hatsuwa', 'sokuji',
]


def daimon_of(path):
    """ファイル名から大問キー（_N5/N4/N3 を除く）を得る。"""
    base = os.path.splitext(os.path.basename(path))[0]
    return re.sub(r'_N[345]$', '', base)


def _txt(d, keys):
    for k in keys:
        v = d.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, list) and v and all(isinstance(x, str) for x in v):
            return '\n'.join(v).strip()
        if isinstance(v, list) and v and all(isinstance(x, dict) for x in v):
            j = '\n'.join(x.get('text') or x.get('body') or x.get('ja') or '' for x in v)
            if j.strip():
                return j.strip()
    return ''


def _choices(d):
    c = d.get('choices') or d.get('audioChoices') or d.get('options')
    if isinstance(c, list):
        out = []
        for x in c:
            if isinstance(x, str):
                out.append(x)
            elif isinstance(x, dict):
                out.append(str(x.get('text') or x.get('ja') or x.get('label') or x))
        return ' / '.join(out)
    return '' if c is None else str(c)


def _answer(d):
    a = d.get('answer')
    if a is None:
        a = d.get('correct')
    if a is None:
        a = d.get('answerIndex')  # 聴解の正解キー
    c = d.get('choices') or d.get('audioChoices') or d.get('options')
    if isinstance(a, int) and isinstance(c, list):
        if 0 <= a < len(c):
            v = c[a]
            return f'#{a}: ' + (v if isinstance(v, str) else str(v))
    return '' if a is None else str(a)


def _risk(d):
    r = str(d.get('uniqRisk', '')).strip().lower()
    if r in ('high', '高', 'red', 'h'):
        return 'high'
    if r in ('mid', 'medium', '中', 'amber', 'm'):
        return 'mid'
    if d.get('ambiguous') is True:
        return 'high'
    if d.get('verified') is False:
        return 'mid'
    return ''


def _note(d, risk):
    n = d.get('uniqNote') or d.get('note')
    if isinstance(n, str) and n.strip():
        return n.strip()
    if d.get('ambiguous') is True:
        return '(暫定) ambiguous=true'
    if d.get('verified') is False:
        return '(暫定) verified=false'
    return ''


def rows_for_item(it):
    """1問→1行以上（passage系は小問ごとに1行）。本文/読み物は必ず出す（無い大問は空欄）。"""
    body = _txt(it, BODY_KEYS)
    subtype = it.get('subtype') or it.get('type') or it.get('qtype') or it.get('format') or it.get('kind') or ''
    subs = it.get('questions')
    base_r = _risk(it)
    out = []
    if isinstance(subs, list) and subs:
        for q in subs:
            if not isinstance(q, dict):
                continue
            r = _risk(q) or base_r
            out.append([
                q.get('id') or it.get('id'), subtype, body,
                _txt(q, Q_KEYS),
                _txt(q, ['underline', 'word']),
                _choices(q), _answer(q), r, _note(q, r) or _note(it, base_r),
                q.get('verified', it.get('verified')),
            ])
    else:
        out.append([
            it.get('id'), subtype,
            body, _txt(it, Q_KEYS),
            _txt(it, ['underline', 'word']),
            _choices(it), _answer(it), base_r, _note(it, base_r),
            it.get('verified'),
        ])
    return out


def collect():
    """(pool, level, daimon) -> list[rows], 出現順の大問リストも保持。"""
    data = {}
    order = {}  # (pool, level) -> [daimon...]
    for f in sorted(glob.glob(SRC + '/**/*.json', recursive=True)):
        rel = os.path.relpath(f, SRC).replace(os.sep, '/')
        try:
            d = json.load(io.open(f, encoding='utf-8'))
        except Exception:
            continue
        items = d.get('items', [])
        if not isinstance(items, list) or not items:
            continue
        pool = d.get('pool') or ('mock' if '/mock/' in rel else 'normal')
        lvl = d.get('level') or (items[0].get('level') if isinstance(items[0], dict) else None)
        if lvl not in LEVELS:
            continue
        dai = daimon_of(f)
        rows = []
        for it in items:
            if isinstance(it, dict):
                rows.extend(rows_for_item(it))
        key = (pool, lvl, dai)
        data.setdefault(key, []).extend(rows)
        order.setdefault((pool, lvl), [])
        if dai not in order[(pool, lvl)]:
            order[(pool, lvl)].append(dai)
    return data, order


def style_header(ws):
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=HEAD)
        cell.alignment = CEN
        cell.border = BORDER


def write_book(pool, lvl, order, data):
    present = order.get((pool, lvl), [])
    if not present:
        return None, 0, 0, 0
    # 本番試験の出題順に並べる（未知の大問は末尾へ）
    dais = [d for d in EXAM_ORDER if d in present]
    dais += [d for d in present if d not in EXAM_ORDER]
    wb = Workbook()
    wb.remove(wb.active)
    n_rows = n_red = n_amber = 0
    for dai in dais:
        rows = data.get((pool, lvl, dai), [])
        ws = wb.create_sheet(JP_NAME.get(dai, dai)[:31])
        ws.append(COLS)
        style_header(ws)
        for r in rows:
            ws.append(['' if x is None else x for x in r])
            rr = ws.max_row
            risk = r[7]
            fill = RED if risk == 'high' else (AMBER if risk == 'mid' else None)
            if fill:
                for c in range(1, len(COLS) + 1):
                    ws.cell(rr, c).fill = PatternFill('solid', fgColor=fill)
                if risk == 'high':
                    n_red += 1
                else:
                    n_amber += 1
            for c in range(1, len(COLS) + 1):
                ws.cell(rr, c).border = BORDER
                ws.cell(rr, c).alignment = WRAP
            n_rows += 1
        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'
    label = '模試' if pool == 'mock' else '通常'
    out = os.path.join(ROOT, f'一意性チェック_{label}_{lvl}.xlsx')
    wb.save(out)
    return out, n_rows, n_red, n_amber


def main():
    data, order = collect()
    print('=== 一意性チェックExcel 生成 ===')
    for pool in ('mock', 'normal'):
        for lvl in LEVELS:
            out, n, red, amber = write_book(pool, lvl, order, data)
            if out:
                print(f'{os.path.basename(out)} | シート{len(order[(pool, lvl)])} 行{n} | 赤{red} 黄{amber}')
    print('出力先:', ROOT)


if __name__ == '__main__':
    main()
