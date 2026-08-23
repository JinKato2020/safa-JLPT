# 「単語×大問カバー率」シートの【語彙単語ブロック】(漢字読み/表記/文脈規定)を現行 content から再計算する。
#   カバー数=その大問が対象にしている「その級の語」の異なり数(=distinct vocabId)。各itemは別語ゆえ 問題数=カバー数。
#   母数=語彙単語ブロックの各大問固有の分母(=シート既存値を踏襲。漢字読み/表記は漢字を持つ語など内在的性質で
#        問題の増減では変わらない)。カバー率=カバー数/母数。
# 用法/言い換え類義/文法ブロックは別ツール(update_usage/synonym/coverage_grammar)。
import json, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

GREEN = PatternFill('solid', fgColor='CDE8D4'); YELLOW = PatternFill('solid', fgColor='FCE7C0'); RED = PatternFill('solid', fgColor='F6C9C4')
def signal(p): return GREEN if p >= 80 else (YELLOW if p >= 60 else RED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
V = json.load(open(os.path.join(ROOT, 'src/data/shared/vocab.json'), encoding='utf-8'))
lvlOf = {v['id']: v['level'] for v in V}
STEM = {'漢字読み': 'kanji_read', '表記': 'orthography', '文脈規定': 'context'}

def covered(stem, lv):
    d = json.load(open(os.path.join(ROOT, f'content/problems/moji_goi/{stem}_{lv}.json'), encoding='utf-8'))
    return len({it['vocabId'] for it in d['items'] if it.get('vocabId') and lvlOf.get(it['vocabId']) == lv})

wb = load_workbook(XLSX)
ws = wb['単語×大問カバー率']
cur_lv = None; done = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and str(a).startswith('■'):
        for lv in ('N5', 'N4', 'N3'):
            if lv in str(a): cur_lv = lv
    b = ws.cell(r, 2).value
    if b and str(b).strip() in STEM and cur_lv:
        label = str(b).strip(); lv = cur_lv
        cov = covered(STEM[label], lv)
        denom = ws.cell(r, 5).value  # 母数は既存値を踏襲(内在的分母)
        try:
            denom = int(denom)
        except (TypeError, ValueError):
            denom = 0
        pct = round(100 * cov / denom) if denom else 0
        ws.cell(r, 3, cov); ws.cell(r, 4, cov)
        c = ws.cell(r, 6, f'{pct}%'); c.fill = signal(pct)
        done.append((lv, label, cov, denom, pct))
wb.save(XLSX)
for d in done:
    print(f'{d[0]} {d[1]}: {d[2]}/{d[3]} = {d[4]}%')
print('saved', XLSX)
