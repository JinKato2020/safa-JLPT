# -*- coding: utf-8 -*-
"""在庫問題数.txt ＋ 模試ストック数.txt を統合して1つのExcelにまとめる。
   レベル×大問で「在庫・未検証・セット数・誤答数内訳」＋「本番出題数・模試換算(何回分)」を横並び。
   使い方: python tools/stock_excel.py
   出力: 問題/在庫・模試ストックまとめ.xlsx"""
import io, os, re, glob, json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
STOCK_TXT = os.path.join(ROOT, 'memory', '在庫問題数.txt')
MOCK_TXT = os.path.join(ROOT, 'memory', '模試ストック数.txt')
OUT = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
LEVELS = ['N5', 'N4', 'N3']

# ── 在庫問題数.txt をパース ──
# section -> [daimon...], (daimon, level) -> {stock, sets, unverified, shuffle, errdist}
stock = {}
sec_of = {}          # daimon -> section
daimon_order = []    # 出現順
shuffle_of = {}      # daimon -> シャッフル
dict_base = []       # 参考: 辞書母数
audited_out = []     # 参考: 在庫外
cur_sec = cur_dai = None
lines = io.open(STOCK_TXT, encoding='utf-8').read().splitlines()
for i, ln in enumerate(lines):
    m = re.match(r'^■\s+(.+?)\s*$', ln)
    if m and '参考' not in m.group(1):
        cur_sec = m.group(1); continue
    m = re.match(r'^【(.+?)】\s*(?:表示選択肢=(\S+)\s*/\s*シャッフル=(\S+))?', ln)
    if m and cur_sec:
        cur_dai = m.group(1)
        if cur_dai not in daimon_order:
            daimon_order.append(cur_dai); sec_of[cur_dai] = cur_sec; shuffle_of[cur_dai] = m.group(3) or ''
        continue
    m = re.match(r'^\s{2}(N[345]):\s*在庫(\d+)問(?:（(\d+)セット）)?\s*/\s*(検証欄なし|未検証(\d+))', ln)
    if m and cur_dai:
        lv = m.group(1)
        rec = {'stock': int(m.group(2)), 'sets': int(m.group(3)) if m.group(3) else None,
               'unverified': None if m.group(4) == '検証欄なし' else int(m.group(5)), 'errdist': ''}
        # 次の数行から誤答数の内訳を拾う
        for j in range(i + 1, min(i + 4, len(lines))):
            e = re.search(r'誤答数\s+(.+?)\s*｜', lines[j])
            if e:
                rec['errdist'] = e.group(1).strip(); break
        stock[(cur_dai, lv)] = rec
        continue
    m = re.match(r'^【(.+?辞書)】\s+(\d+)件（(.+?)）', ln)
    if m:
        dict_base.append((m.group(1), int(m.group(2)), m.group(3))); continue
    m = re.match(r'^【(.+?)】\s+監査に落ちて在庫外\s+(\d+)問', ln)
    if m:
        audited_out.append((m.group(1), int(m.group(2)))); continue

# ── 模試ストック数.txt をパース（本番出題数）──
honban = {}  # (daimon, level) -> 本番数(int)
for ln in io.open(MOCK_TXT, encoding='utf-8').read().splitlines():
    m = re.match(r'^\s{2}(\S+)\s+(N[345])\s+(\d+)\s+(\d+)\s+(\S+)\s*$', ln)
    if m:
        honban[(m.group(1), m.group(2))] = int(m.group(4))

# ── 漢字読み/表記の在庫＝「出題級で絞った"問題数"(eligible・拡張後の実数)」で上書き（2026-08-29）──
# 生カウントは他級の漢字を含む語も数えて過大。出題級=testLevel=MAX(語彙級,全漢字級)。class=kanjiのみ漢字読み対象。
# ★語数でなく問題数で数える＝1語に複数問(学習拡張)がある場合その実数を反映(例 漢字読みN5=146語→410問)。mock/配下は除外(通常在庫でない)。
_V = {x['id']: x for x in json.load(io.open(os.path.join(ROOT, 'src', 'data', 'shared', 'vocab.json'), encoding='utf-8'))}
_VKC = json.load(io.open(os.path.join(ROOT, 'src', 'data', 'shared', 'vocabKanjiClass.json'), encoding='utf-8'))['items']
def _bank_vid_list(daimon):
    out = []
    for p in glob.glob(os.path.join(ROOT, 'content', 'problems', 'moji_goi', daimon + '_*.json')):
        for it in json.load(io.open(p, encoding='utf-8')).get('items', []):
            if it.get('vocabId'): out.append(it['vocabId'])
    return out
_KR = _bank_vid_list('kanji_read'); _OG = _bank_vid_list('orthography')
def _elig_kr(v, lv):
    r = _VKC.get(v, {}); return r.get('class') == 'kanji' and r.get('testLevel') == lv
def _elig_og(v, lv):
    r = _VKC.get(v, {}); return (r.get('class') == 'kanji' and r.get('testLevel') == lv) or (r.get('class') == 'katakana' and _V.get(v, {}).get('level') == lv)
for lv in LEVELS:
    kre = sum(1 for v in _KR if v in _V and _elig_kr(v, lv))   # 出題級で絞った"問題数"(拡張後の実数)
    oge = sum(1 for v in _OG if v in _V and _elig_og(v, lv))
    if ('漢字読み', lv) in stock: stock[('漢字読み', lv)]['stock'] = kre
    if ('表記', lv) in stock: stock[('表記', lv)]['stock'] = oge

# ── 統合行を作る ──
rows = []  # (section, daimon, level, stock, sets, unverified, honban, kaisu, errdist, shuffle)
for dai in daimon_order:
    for lv in LEVELS:
        if (dai, lv) not in stock:
            continue
        r = stock[(dai, lv)]
        hb = honban.get((dai, lv))
        kaisu = (r['stock'] // hb) if hb else None
        rows.append((sec_of[dai], dai, lv, r['stock'], r['sets'], r['unverified'], hb, kaisu, r['errdist'], shuffle_of.get(dai, '')))

# 律速（セクション別・級ごと）＝そのくくりで最小の模試換算
def min_kaisu(pred):
    out = {}
    for (_, _, lv, _, _, _, hb, k, _, _) in rows:
        if hb and k is not None and pred(_, lv):
            out[lv] = min(out.get(lv, 10**9), k)
    return out
sections = ['文字・語彙', '文法', '読解', '聴解']
sec_ritsu = {sec: {lv: min((k for (s, d, l, st, se, uv, hb, k, ed, sh) in rows if s == sec and l == lv and hb and k is not None), default=None) for lv in LEVELS} for sec in sections}
full_mock = {lv: min((k for (s, d, l, st, se, uv, hb, k, ed, sh) in rows if l == lv and hb and k is not None), default=None) for lv in LEVELS}

# ── Excel 生成 ──
NAVY = '1F3B57'; HEAD = '2E5A88'; SUB = 'DDE7F2'
FILL_SEC = {'文字・語彙': 'FDF2E3', '文法': 'EAF3EA', '読解': 'EAF0FA', '聴解': 'F3EAF3'}
thin = Side(style='thin', color='C9D3DE')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=HEAD)
        cell.alignment = CEN; cell.border = BORDER

def kaisu_fill(k):
    if k is None: return None
    if k <= 5: return 'F6C9C4'      # 赤=不足
    if k <= 15: return 'FCE7C0'     # 橙=要増産
    return 'CDE8D4'                 # 緑=十分

# 既存ブックがあれば読み込み、自分の3シートだけ差し替える（習得シート等の手作りシートを保全）。
_OWN = ('大問別まとめ', '模試ストック換算', '参考')
if os.path.exists(OUT):
    wb = load_workbook(OUT)
    for _nm in _OWN:
        if _nm in wb.sheetnames: del wb[_nm]
    ws = wb.create_sheet('大問別まとめ', 0)  # 在庫まとめを先頭に
else:
    wb = Workbook()
    ws = wb.active; ws.title = '大問別まとめ'
# 学習の割増し倍数（通常問題数を語あたり何問に増やしたか）・模試プールの問題数・メモ。
# 現状=漢字読みのみ×3割増し／模試プールは漢字読み(104/90/80)・表記(80/60/60)の2大問に新設。他大問は未実施=×1。
MULT = {('漢字読み', 'N5'): '×3'}
MOCK = {('漢字読み', 'N5'): 104, ('漢字読み', 'N4'): 90, ('漢字読み', 'N3'): 80,
        ('表記', 'N5'): 80, ('表記', 'N4'): 60, ('表記', 'N3'): 60,
        ('文脈規定', 'N5'): 100, ('文脈規定', 'N4'): 100, ('文脈規定', 'N3'): 110,
        ('言い換え類義', 'N5'): 50, ('言い換え類義', 'N4'): 50, ('言い換え類義', 'N3'): 50,
        ('用法', 'N4'): 50, ('用法', 'N3'): 50}
MEMO = {('漢字読み', 'N5'): 'N5の質の高い漢字語（数字・接尾辞を除く）は実質104語＝N5は漢字自体が少ないため上限。模試プールは120でなく104問が自然な上限。',
        ('表記', 'N5'): '漢字読みN5の104×2の例題を流用'}
cols = ['セクション', '大問', 'レベル', '在庫問題数', '割増し倍数', '模試問題数', 'メモ', 'セット数', '本番出題数', '模試換算(何回分)', '誤答数の内訳', 'シャッフル']
ws.append(cols); style_header(ws, len(cols))
for (sec, dai, lv, st, se, uv, hb, k, ed, sh) in rows:
    mult = MULT.get((dai, lv), '×1')
    mock = MOCK.get((dai, lv), '—')
    memo = MEMO.get((dai, lv), '')
    ws.append([sec, dai, lv, st, mult, mock, memo, (se if se is not None else '—'),
               (hb if hb is not None else '—'),
               ('—' if k is None else k), (ed or '—'), sh])
    r = ws.max_row
    ws.cell(r, 1).fill = PatternFill('solid', fgColor=FILL_SEC.get(sec, 'FFFFFF'))
    for c in range(1, len(cols) + 1):
        cell = ws.cell(r, c); cell.border = BORDER
        cell.alignment = LEFT if c in (1, 2, 7, 11) else CEN  # セクション/大問/メモ/誤答数内訳=左寄せ
    if mult != '×1':
        ws.cell(r, 5).font = Font(bold=True, color='1F7A3D')  # 割増し実施は緑太字
    fc = kaisu_fill(k)
    if fc:
        kc = ws.cell(r, 10); kc.fill = PatternFill('solid', fgColor=fc); kc.font = Font(bold=True)
ws.freeze_panes = 'D2'
widths = [12, 16, 7, 12, 9, 10, 46, 9, 11, 15, 26, 12]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.cell(1, 6).comment = None
ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

# Sheet2: 模試ストック（律速・フル模試）
ws2 = wb.create_sheet('模試ストック換算', 1)
ws2.append(['◆ 模試換算 = 在庫 ÷ 本番出題数 = その大問だけで組めるフル模試の回数（floor）'])
ws2.cell(1, 1).font = Font(bold=True, size=12, color=NAVY)
ws2.append([])
ws2.append(['セクション別 律速（そのセクションで最も少ない大問の回数＝そのセクションを何回組めるか）'])
ws2.cell(3, 1).font = Font(bold=True, color=NAVY)
hdr = ['セクション'] + LEVELS
ws2.append(hdr); hr = ws2.max_row; style_header(ws2, len(hdr), hr)
for sec in sections:
    ws2.append([sec] + [sec_ritsu[sec][lv] if sec_ritsu[sec][lv] is not None else '—' for lv in LEVELS])
    r = ws2.max_row
    ws2.cell(r, 1).fill = PatternFill('solid', fgColor=FILL_SEC.get(sec, 'FFFFFF'))
    for c in range(1, len(hdr) + 1):
        cell = ws2.cell(r, c); cell.border = BORDER; cell.alignment = CEN if c > 1 else LEFT
        if c > 1:
            fc = kaisu_fill(sec_ritsu[sec][LEVELS[c - 2]])
            if fc: cell.fill = PatternFill('solid', fgColor=fc)
ws2.append([])
ws2.append(['級ごとの「フル模試」完成可能回数（全大問の最小＝律速大問で決まる）'])
ws2.cell(ws2.max_row, 1).font = Font(bold=True, color=NAVY)
ws2.append(hdr); hr2 = ws2.max_row; style_header(ws2, len(hdr), hr2)
ws2.append(['フル模試'] + [full_mock[lv] if full_mock[lv] is not None else '—' for lv in LEVELS])
r = ws2.max_row
for c in range(1, len(hdr) + 1):
    cell = ws2.cell(r, c); cell.border = BORDER; cell.alignment = CEN if c > 1 else LEFT
    if c > 1:
        fc = kaisu_fill(full_mock[LEVELS[c - 2]])
        if fc: cell.fill = PatternFill('solid', fgColor=fc); cell.font = Font(bold=True)
ws2.append([]); ws2.append(['凡例：赤=5回以下(不足) / 橙=15回以下(要増産) / 緑=16回以上(十分)'])
ws2.cell(ws2.max_row, 1).font = Font(italic=True, color='7A756C')
for i, w in enumerate([14, 8, 8, 8], 1): ws2.column_dimensions[get_column_letter(i)].width = w

# 参考シートは廃止（ユーザー指示 2026-08-30）。_OWN に残すことで既存の参考シートは再生成時に削除される。
total_stock = sum(st for (_, _, _, st, _, _, _, _, _, _) in rows)  # 集計ログ用にのみ計算

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('wrote', OUT)
print('rows:', len(rows), '/ total_stock:', total_stock)
print('full_mock:', full_mock)
