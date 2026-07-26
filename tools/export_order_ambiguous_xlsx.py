# 「文の組み立て」(並べ替え)のうち、一意性監査で ambiguous:true が付いて出題から外れている問題をExcelに出す
#   python tools\export_order_ambiguous_xlsx.py
# 出典 = app\content\problems\bunpou\order_{N5,N4,N3}.json（＝ビルドに乗っている正本）
import json, os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'app', 'content', 'problems', 'bunpou')
GRAMMAR = os.path.join(ROOT, 'app', 'src', 'data', 'shared', 'grammar.json')
XLSX = os.path.join(ROOT, '文の組み立て_除外分.xlsx')

gmap = {g['id']: g for g in json.load(open(GRAMMAR, encoding='utf-8'))}

wb = Workbook(); wb.remove(wb.active)
head = ['問題ID', '問題文（★の位置が答え）', '設問文', '選択肢1', '選択肢2', '選択肢3', '選択肢4',
        '★の答え', '答えは選択肢の何番目', 'pointId', '文法項目', '項目の級']
summary = []

for lv in ['N5', 'N4', 'N3']:
    d = json.load(open(os.path.join(SRC, f'order_{lv}.json'), encoding='utf-8'))
    amb = [i for i in d['items'] if i.get('ambiguous')]
    ws = wb.create_sheet(lv)
    ws.append(head)
    for it in amb:
        ch = (it.get('choices') or []) + [''] * 4
        ans = it.get('answer', '')
        g = gmap.get(it.get('pointId'))
        # 選択肢の中で答えが何番目か（人が見比べる時の目印。データ上の正解は answer 文字列そのもの）
        pos = (it['choices'].index(ans) + 1) if ans in (it.get('choices') or []) else ''
        ws.append([it['id'], it.get('stem', ''), it.get('question', ''),
                   ch[0], ch[1], ch[2], ch[3], ans, pos,
                   it.get('pointId', ''), g['point'] if g else '', g['level'] if g else ''])
    for col, w in zip('ABCDEFGHIJKL', [12, 60, 30, 18, 18, 18, 18, 18, 9, 10, 22, 8]):
        ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical='top')
        row[2].alignment = Alignment(wrap_text=True, vertical='top')
    ws.freeze_panes = 'A2'
    summary.append((lv, len(d['items']), len(amb), len(d['items']) - len(amb)))
    print(f"{lv}: 全{len(d['items'])}問 中 除外{len(amb)}問（実際に出るのは{len(d['items']) - len(amb)}問）")

ws = wb.create_sheet('集計', 0)
ws.append(['級', '全問', '除外（ambiguous）', '実際に出題', '除外率'])
for lv, tot, a, ok in summary:
    ws.append([lv, tot, a, ok, f'{a / tot * 100:.0f}%'])
T, A, O = sum(r[1] for r in summary), sum(r[2] for r in summary), sum(r[3] for r in summary)
ws.append(['合計', T, A, O, f'{A / T * 100:.0f}%'])
ws.append([])
ws.append(['除外の理由 = 一意性監査(2026-07-10)で「別の語順でも自然な文になり、★の答えが複数ありうる」と判定されたもの'])
ws.append(['アプリは ambiguous:true の問題を出題プールから外している（app/src/data/daimon.ts）'])
for col, w in zip('ABCDE', [8, 10, 18, 12, 10]):
    ws.column_dimensions[col].width = w
for c in ws[1]:
    c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')

wb.save(XLSX)
print(f"\n除外 合計 {A}問 / 全{T}問")
print(f"Excel: {XLSX}")
