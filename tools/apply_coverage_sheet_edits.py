# 単語×大問カバー率シートへの一括反映(この会話の確定分):
#  ①H6/H7(N5 漢字読み/表記)に「かな語は母数外」の説明を追記
#  ②N5 用法行をグレーアウト(公式試験に用法大問が無い=対象外)。※N4/N3 用法は実在ゆえ据置。
# Excelを閉じてから: python tools/apply_coverage_sheet_edits.py
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
GRAY = PatternFill('solid', fgColor='D9D9D9')
EXPL = ('漢字読み・表記＝「その漢字をどう読む／どう書く」を問う大問。漢字を含む語でしか作れません。'
        'N5の723語のうち、かな・カタカナだけの語（例：あります・きれい・テレビ・ゆっくり）約160語には'
        '漢字形が無いので、この2大問の母数から自動で外れ、563語になります。')

wb = load_workbook(XLSX)
ws = wb['単語×大問カバー率']
for r in (6, 7):  # H6=N5漢字読み, H7=N5表記
    c = ws.cell(r, 8, EXPL)
    c.alignment = Alignment(wrap_text=True, vertical='top')

cur = None; grayed = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and str(a).startswith('■'):
        for lv in ('N5', 'N4', 'N3'):
            if lv in str(a):
                cur = lv
    b = ws.cell(r, 2).value
    if b and str(b).strip() == '用法' and cur == 'N5':
        for cc in range(1, 9):
            ws.cell(r, cc).fill = GRAY
        ws.cell(r, 6, '対象外'); ws.cell(r, 6).font = Font(color='808080')
        ws.cell(r, 8, 'N5の文字・語彙に「用法」大問は無い(公式試験対象外)。アプリ予想得点も既にN5用法を除外済み。').alignment = Alignment(wrap_text=True, vertical='top')
        grayed.append(r)

wb.save(XLSX)
print('OK  H6/H7追記 + N5用法グレー行=', grayed, '->', XLSX)
