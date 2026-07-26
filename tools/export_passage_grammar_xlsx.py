# アプリに入っている「文章の文法」を、そのままExcelに出す
#   python tools\export_passage_grammar_xlsx.py
# 出典 = app\content\problems\bunpou\passage_grammar_{N5,N4,N3}.json（＝ビルドに乗っている正本）
import json, os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'app', 'content', 'problems', 'bunpou')
GRAMMAR = os.path.join(ROOT, 'app', 'src', 'data', 'shared', 'grammar.json')
XLSX = os.path.join(ROOT, '文章の文法_ビルド収録分.xlsx')

gmap = {g['id']: g for g in json.load(open(GRAMMAR, encoding='utf-8'))}

wb = Workbook(); wb.remove(wb.active)
head = ['セットID', '空所', '本文（その空所を含む方）', '選択肢1', '選択肢2', '選択肢3', '選択肢4',
        '正解(1-4)', 'pointId', '文法項目', '項目の級']
summary = []

for lv in ['N5', 'N4', 'N3']:
    d = json.load(open(os.path.join(SRC, f'passage_grammar_{lv}.json'), encoding='utf-8'))
    ws = wb.create_sheet(lv)
    ws.append(head)
    nq = 0
    for it in d['items']:
        bodies = [p['body'] for p in it['passages']]
        for q in it['questions']:
            bn = q.get('blankNo')
            # 空所が何本目の本文にあるか探す（本文1本のセットもある）
            body = next((b for b in bodies if f'【{bn}】' in b), bodies[0])
            g = gmap.get(q.get('pointId'))
            ai = q.get('answerIndex')
            ch = (q.get('choices') or []) + [''] * 4
            ws.append([it['id'], bn, body, ch[0], ch[1], ch[2], ch[3],
                       (ai + 1) if isinstance(ai, int) else '',
                       q.get('pointId', ''), g['point'] if g else '', g['level'] if g else ''])
            nq += 1
    for col, w in zip('ABCDEFGHIJK', [12, 6, 90, 16, 16, 16, 16, 9, 10, 22, 8]):
        ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical='top')
    ws.freeze_panes = 'A2'
    summary.append((lv, d.get('schema'), d.get('languages'), len(d['items']), nq))
    print(f"{lv}: {len(d['items'])}セット / {nq}問 / languages={d.get('languages')}")

ws = wb.create_sheet('集計', 0)
ws.append(['級', 'schema', 'languages', 'セット数', '問題数'])
for r in summary:
    ws.append([r[0], r[1], ','.join(r[2] or []), r[3], r[4]])
ws.append(['合計', '', '', sum(r[3] for r in summary), sum(r[4] for r in summary)])
for col, w in zip('ABCDE', [8, 8, 12, 10, 10]):
    ws.column_dimensions[col].width = w
for c in ws[1]:
    c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDDDDD')

wb.save(XLSX)
print(f"\n合計 {sum(r[3] for r in summary)}セット / {sum(r[4] for r in summary)}問")
print(f"Excel: {XLSX}")
