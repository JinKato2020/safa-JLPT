# 「単語×大問カバー率」シートの【文法単語ブロック】(文法形式/語の並べ替え/文章の文法)を
# 現行 content から再計算して更新する。※このシートは元々2026-08-19の手集計スナップショットで
# 自動更新ツールが無かった＝文章の文法の増作がカバー率に反映されない問題を塞ぐ。
# 母数=レベル内の文法点数(指標対象外 n5-g-92 を除外＝アプリのカバー率定義 selectors.coverageBars と統一)。
# 語彙単語/漢字単語ブロックは vocabId 実リンク集計(別ロジック)ゆえ本ツールの対象外(据置)。
import json, os, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 信号色(カバー率%セル)= アプリ在庫シートの既存配色に一致
#   緑 ≥80% / 黄 60-79% / 赤 <60%
GREEN = PatternFill('solid', fgColor='CDE8D4')
YELLOW = PatternFill('solid', fgColor='FCE7C0')
RED = PatternFill('solid', fgColor='F6C9C4')


def signal(pct):
    return GREEN if pct >= 80 else (YELLOW if pct >= 60 else RED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
SRC = os.path.join(ROOT, 'content', 'problems', 'bunpou')
EXCL = {'n5-g-92'}

g = json.load(open(os.path.join(ROOT, 'src/data/shared/grammar.json'), encoding='utf-8'))
lvl = {x['id']: x['level'] for x in g}
TOT = {lv: len({x['id'] for x in g if x['level'] == lv and x['id'] not in EXCL})
       for lv in ['N5', 'N4', 'N3']}

# 大問ラベル(col B) → content ファイル接頭辞
DAIMON = {'文法形式(穴埋め)': 'grammar_form', '語の並べ替え': 'order', '文章の文法': 'passage_grammar'}


def stats(prefix, lv):
    d = json.load(open(os.path.join(SRC, f'{prefix}_{lv}.json'), encoding='utf-8'))
    items = d['items'] if isinstance(d, dict) and 'items' in d else d
    pids = []
    for it in items:
        for q in it.get('questions', [it]):
            pid = (q or it).get('pointId')
            if pid:
                pids.append(pid)
    nq = len(pids)
    cov = len({p for p in pids if lvl.get(p) == lv and p not in EXCL})
    return nq, cov


wb = load_workbook(XLSX)
ws = wb['② カバー率' if '② カバー率' in wb.sheetnames else '単語×大問カバー率']
changed = []
cur_lv = None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and str(a).startswith('■'):
        m = re.search(r'N[345]', str(a))
        cur_lv = m.group(0) if m else None
    if a and str(a).startswith('文法単語'):
        ws.cell(r, 1, f'文法単語({TOT[cur_lv]})')  # 母数ラベルも更新
    b = ws.cell(r, 2).value
    if not b or cur_lv is None:
        continue
    daimon = str(b).strip()
    if daimon in DAIMON:
        nq, cov = stats(DAIMON[daimon], cur_lv)
        tot = TOT[cur_lv]
        pctv = f'{round(cov / tot * 100)}%'
        old = [ws.cell(r, c).value for c in (3, 4, 5, 6)]
        ws.cell(r, 3, nq); ws.cell(r, 4, cov); ws.cell(r, 5, tot); ws.cell(r, 6, pctv)
        ws.cell(r, 6).fill = signal(round(cov / tot * 100))  # 信号色を再適用
        new = [nq, cov, tot, pctv]
        if old != new:
            changed.append(f'{cur_lv} {daimon}: {old} -> {new}')

# 文法ブロック以外も含め、%セル(col6)の信号色を現値に合わせて総ざらい(古い色を是正)
recolored = 0
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 6).value
    if isinstance(v, str) and v.strip().endswith('%'):
        try:
            pct = int(v.strip().rstrip('%'))
        except ValueError:
            continue
        ws.cell(r, 6).fill = signal(pct)
        recolored += 1
print('信号色 再適用セル数:', recolored)

wb.save(XLSX)
print('WROTE', XLSX)
print('母数(除外後):', TOT)
for c in changed:
    print(' ', c)
