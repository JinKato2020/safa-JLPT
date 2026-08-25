# -*- coding: utf-8 -*-
"""新規作問(usage_new_result_*.json)を用法N4へ反映する。
  1) 採用(adopt)問を検証: 誤答3・repl一意(P1)・type2種以上(P2)・type語彙正当・vocabId=N4かつ未カバー
  2) id連番(N4-V-Y-最大+1〜)を採番し usage_N4.json に追加
  3) サイドカー usageDistractorTags.json に tags 追記(certainty含む)
  4) borderline誤答をオレンジ着色した確認用Excelを生成
検証NGは適用せず needs_manual に退避(データを壊さない)。--dry で試算のみ。
"""
import json, io, os, re, sys, argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SP = r'C:\Users\jwpsa\AppData\Local\Temp\claude\c--Users-jwpsa-Documents-desktop-claude-JLPT---\21dfc579-d8da-4b72-ae72-cc54a810863c\scratchpad'
JSON_N4 = 'content/problems/moji_goi/usage_N4.json'
VOCAB   = 'src/data/shared/vocab.json'
SIDE    = 'src/data/shared/usageDistractorTags.json'
XLSX    = '用法N4_新規作問_確認用.xlsx'
TYPEVOCAB = {'自他', '別義', '近接', '選択', 'コロケ', '対義', '呼応', '授受'}
ORANGE = PatternFill('solid', fgColor='FFC000')
HEAD   = PatternFill('solid', fgColor='DDDDDD')


def q_text(stem):
    return f'「{stem}」の使（つか）い方（かた）として最（もっと）もよいものはどれですか。'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--batches', type=int, default=10)
    ap.add_argument('--sp', default=SP, help='結果JSON(usage_new_result_*.json)の置き場。既定=現行スクラッチパッド')
    a = ap.parse_args()
    sp = a.sp

    vocab = {v['id']: v for v in json.load(open(VOCAB, encoding='utf-8'))}
    d = json.load(open(JSON_N4, encoding='utf-8'))
    items = d['items']
    covered = {it.get('vocabId') for it in items if it.get('vocabId')}
    # 「1語彙id=1問(全大問通算)」の番人に合わせ、他の用法ファイル(N3/N5)で既に使われたvocabIdも除外
    for lv in ('N3', 'N5'):
        fp = f'content/problems/moji_goi/usage_{lv}.json'
        if os.path.exists(fp):
            covered |= {it.get('vocabId') for it in json.load(open(fp, encoding='utf-8'))['items']
                        if it.get('vocabId')}
    mx = max(int(re.search(r'(\d+)$', it['id']).group(1))
             for it in items if it['id'].startswith('N4-V-Y-'))
    side = json.load(open(SIDE, encoding='utf-8'))

    # 結果収集
    results = []
    for i in range(1, a.batches + 1):
        rp = os.path.join(sp, f'usage_new_result_{i}.json')
        if not os.path.exists(rp):
            print(f'!! 結果ファイル無し: {rp}')
            continue
        try:
            results.extend(json.load(open(rp, encoding='utf-8')))
        except Exception as e:
            print(f'!! 破損: {rp}: {e}')

    applied, skipped_by_agent, needs_manual, rows = [], [], [], []
    seen_vocab = set()
    nxt = mx
    for r in results:
        vid = r.get('vocabId')
        if not r.get('adopt'):
            skipped_by_agent.append((vid, r.get('reason', '')))
            continue
        reason = None
        dists = r.get('distractors', [])
        if vid not in vocab or vocab[vid].get('level') != 'N4':
            reason = f'vocabId不正/非N4: {vid}'
        elif vid in covered or vid in seen_vocab:
            reason = f'既カバー/重複vocabId: {vid}'
        elif len(dists) != 3:
            reason = f'誤答数≠3: {len(dists)}'
        else:
            repls = [x.get('repl', '') for x in dists]
            types = [x.get('type', '') for x in dists]
            sents = [x.get('sent', '') for x in dists]
            if not all(r.get('stem') and r.get('answer')):
                reason = 'stem/answer欠落'
            elif any(t not in TYPEVOCAB for t in types):
                reason = f'type語彙外: {types}'
            elif len(set(repls)) < 3:
                reason = f'repl重複(P1): {repls}'
            elif len(set(types)) < 2 and types[0] not in ('選択', '呼応'):
                reason = f'型が全同一(P2)かつ非例外型: {types}'
            elif any(not s for s in sents):
                reason = 'sent空'
            elif r['answer'] in sents:
                reason = 'answerが誤答に混入'
        if reason:
            needs_manual.append((vid, reason))
            continue

        seen_vocab.add(vid)
        nxt += 1
        iid = f'N4-V-Y-{nxt:04d}'
        sents = [x['sent'] for x in dists]
        it = {'id': iid, 'stem': r['stem'], 'question': q_text(r['stem']),
              'answer': r['answer'], 'choices': [r['answer']] + sents,
              'i18n': {}, 'verified': True, 'vocabId': vid}
        if not a.dry:
            items.append(it)
            side['tags'][iid] = [{'repl': x['repl'], 'type': x['type'],
                                  'certainty': x.get('certainty', 'clear')} for x in dists]
            # 選択制限型/否定呼応型など、公式が認める単一殺し方の良問は monoTypeAllow に明示登録(番人P2の例外)。
            if len({x['type'] for x in dists}) == 1:
                side.setdefault('monoTypeAllow', [])
                if iid not in side['monoTypeAllow']:
                    side['monoTypeAllow'].append(iid)
        applied.append(iid)
        rows.append({'id': iid, 'stem': r['stem'], 'answer': r['answer'],
                     'dists': dists})

    if not a.dry:
        json.dump(d, io.open(JSON_N4, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
        json.dump(side, io.open(SIDE, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

    # 確認用Excel
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '用法N4 新規作問'
    hdr = ['ID', '語', '○ 正解文', '誤答1', '型1', '誤答2', '型2', '誤答3', '型3']
    ws.append(hdr)
    for c in ws[1]: c.fill = HEAD; c.font = Font(bold=True)
    ws.append(['凡例: オレンジ=一意性あやしい(要目視)', *[''] * 8])
    bcnt = 0
    for row in rows:
        line = [row['id'], row['stem'], row['answer']]
        for x in row['dists']:
            line += [x['sent'], f"{x['type']}→{x['repl']}"]
        ws.append(line)
        ri = ws.max_row
        for j, x in enumerate(row['dists']):
            if x.get('certainty') == 'borderline':
                ws.cell(ri, 4 + j * 2).fill = ORANGE; bcnt += 1
    for col, w in zip('ABCDEFGHI', [13, 10, 42, 36, 14, 36, 14, 36, 14]):
        ws.column_dimensions[col].width = w
    for r_ in ws.iter_rows():
        for c in r_: c.alignment = Alignment(vertical='top', wrap_text=True)
    if not a.dry:
        wb.save(XLSX)

    print(f'{"[DRY] " if a.dry else ""}applied {len(applied)} / needs_manual {len(needs_manual)} '
          f'/ agent_skip {len(skipped_by_agent)} / borderline誤答 {bcnt}')
    if applied:
        print(f'  id: {applied[0]} 〜 {applied[-1]}')
    if needs_manual:
        print('--- needs_manual ---')
        for x in needs_manual: print('  ', x)
    print('XLSX:', os.path.abspath(XLSX))


if __name__ == '__main__':
    main()
