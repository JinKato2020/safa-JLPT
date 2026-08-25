# 用法(usage)大問が語彙idを「バランスよく」カバーしているかを記録する。
#   記録先(正本)= memory/在庫・模試ストックまとめ.xlsx の専用シート「⑤ 用法カバー×バランス」(毎回作り直し=冪等)。
#   併せて番人の基準 = src/data/shared/usageCoverage.json を --set-baseline で更新。
#
# 測る本質パラメータ = ①breadth(何語をカバーしたか=coverage%) ②backlog(未カバー語=今後の作問対象)
#   ③concentration(1語に問題を集中させていないか) ④level-fit(その級の語を出題しているか=下級語へ流出していないか)
#   ⑤級跨ぎ重複(同語をN4/N3両方で出題=冗長)。最終目標=可能な限り語彙をカバー(breadth%→100%)。
#   ※意味「分野(category)」の均等さは指標にしない(breadthを100%へ寄せれば分野は自動で埋まる=本質の邪魔)。
#
# 使い方:
#   python tools/usage_coverage_report.py                 # 集計してExcelを更新(基準は変更しない)
#   python tools/usage_coverage_report.py --set-baseline  # 上記 + 番人基準(usageCoverage.json)を現状で更新
import json, os, sys
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
SHEET = '⑤ 用法カバー×バランス'
BASELINE = os.path.join(ROOT, 'src/data/shared/usageCoverage.json')
LEVELS = ('N4', 'N3')            # 用法大問が存在する級(N5に用法は無い)
RANK = {'N5': 0, 'N4': 1, 'N3': 2}

GREEN = PatternFill('solid', fgColor='CDE8D4'); YELLOW = PatternFill('solid', fgColor='FCE7C0'); RED = PatternFill('solid', fgColor='F6C9C4')
HEAD = PatternFill('solid', fgColor='D9D9D9'); BOLD = Font(bold=True)
def signal(p): return GREEN if p >= 80 else (YELLOW if p >= 60 else RED)

V = json.load(open(os.path.join(ROOT, 'src/data/shared/vocab.json'), encoding='utf-8'))
lvlOf = {v['id']: v['level'] for v in V}
vocTot = {lv: sum(1 for v in V if v['level'] == lv) for lv in ('N5', 'N4', 'N3')}

# --- 用法データを読む ---
items_by_daimon = {}
for lv in LEVELS:
    d = json.load(open(os.path.join(ROOT, f'content/problems/moji_goi/usage_{lv}.json'), encoding='utf-8'))
    items_by_daimon[lv] = d['items'] if isinstance(d, dict) else d

# 級跨ぎ重複(同じ語をN4とN3の両大問で出題=冗長カバー)を記録用に集計
daimon_of_vid = defaultdict(set)
for dlv in LEVELS:
    for it in items_by_daimon[dlv]:
        if it.get('vocabId'):
            daimon_of_vid[it['vocabId']].add(dlv)
cross_dup = sorted(vid for vid, ds in daimon_of_vid.items() if len(ds) > 1)

# breadth: 語の「本来の級」でカバー集計(N3大問がN4語を出題→N4カバーに計上)
covByLvl = {lv: set() for lv in ('N5', 'N4', 'N3')}
daimon_stat = {}
unlinked_stems = []
for dlv in LEVELS:
    conc = Counter(); fit = Counter(); unl = 0
    for it in items_by_daimon[dlv]:
        vid = it.get('vocabId')
        if not vid:
            unl += 1; s = it.get('stem')
            if s: unlinked_stems.append(s)
            continue
        wl = lvlOf.get(vid)
        conc[vid] += 1
        fit[wl] += 1
        if wl in covByLvl:
            covByLvl[wl].add(vid)
    dups = {k: n for k, n in conc.items() if n > 1}
    daimon_stat[dlv] = {
        'items': len(items_by_daimon[dlv]),
        'linked': sum(conc.values()),
        'unlinked': unl,
        'dupWords': len(dups),
        'maxConc': max(conc.values()) if conc else 0,
        'fit': dict(fit),          # 出題語の本来の級の内訳
    }

# per-level breadth + backlog
level_report = {}
for lv in LEVELS:
    cov = len(covByLvl[lv]); tot = vocTot[lv]
    level_report[lv] = {
        'cov': cov, 'tot': tot,
        'pct': round(100 * cov / tot) if tot else 0,
        'backlog': tot - cov,
    }

# ---------------- Excel 記録(専用シートを作り直し) ----------------
wb = load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)
ws.column_dimensions['A'].width = 16
for c in 'BCDEFGH':
    ws.column_dimensions[c].width = 12
ws.column_dimensions['I'].width = 66

r = 1
ws.cell(r, 1, '用法(語の使い方) × 語彙idカバー×バランス').font = Font(bold=True, size=13); r += 2
ws.cell(r, 1, '最終目標＝可能な限り語彙をカバー(breadth%を100%へ)。番人=src/data/usageCoverage.test.ts が'
              '「①未紐づけ0(測定可能) ②大問内で1語1問まで(集中させず広く) ③カバー数は後退させない(ラチェット)」を担保。').alignment = Alignment(wrap_text=True); r += 1
ws.cell(r, 1, f'級跨ぎ重複(同語をN4とN3の両大問で出題=冗長)＝{len(cross_dup)}語。'
              '冗長は禁止せず記録のみ(復習として妥当な場合あり)＝新規作問は未カバー語を優先すると breadth が伸びる。').alignment = Alignment(wrap_text=True); r += 2

# サマリ表(本質パラメータのみ)
ws.cell(r, 1, '■ 級別サマリ').font = BOLD; r += 1
hdr = ['級', '問題数', 'リンク', '未紐づけ', 'カバー語', '母数', 'breadth%', '未カバー(backlog)']
for i, h in enumerate(hdr):
    c = ws.cell(r, i + 1, h); c.font = BOLD; c.fill = HEAD; c.alignment = Alignment(horizontal='center')
r += 1
for lv in LEVELS:
    ds = daimon_stat[lv]; lr = level_report[lv]
    fit = ds['fit']; fitnote = '級内=%d' % fit.get(lv, 0)
    leak = {k: v for k, v in fit.items() if k != lv}
    if leak:
        fitnote += ' / 下級流出=' + ', '.join(f'{k}:{v}' for k, v in sorted(leak.items(), key=lambda x: RANK[x[0]]))
    vals = [lv, ds['items'], ds['linked'], ds['unlinked'], lr['cov'], lr['tot'], f"{lr['pct']}%", lr['backlog']]
    for i, v in enumerate(vals):
        cc = ws.cell(r, i + 1, v)
        if i == 6:
            cc.fill = signal(lr['pct'])
    ws.cell(r, 9, f"集中: 重複語{ds['dupWords']}・最大{ds['maxConc']}問/語 ｜ level-fit: {fitnote}").alignment = Alignment(wrap_text=True)
    r += 1

wb.save(XLSX)

# ---------------- 番人基準(usageCoverage.json) ----------------
if '--set-baseline' in sys.argv:
    # 未紐づけstemは「vocabマスタ未収録で測定外」の許容リスト(自己回復=次回再生成で最新化)
    allow = sorted(set(unlinked_stems))
    payload = {
        'note': '用法(usage)のvocabIdカバー×バランス基準。tools/usage_coverage_report.py --set-baseline で更新。',
        'goal': '可能な限り語彙をカバー(breadth%→100%)。番人はカバー数の後退と集中(1語2問以上)を禁じる。',
        'baseline': {lv: {'covered': level_report[lv]['cov'], 'total': level_report[lv]['tot'], 'pct': level_report[lv]['pct']} for lv in LEVELS},
        'unlinkedAllowlist': allow,
    }
    json.dump(payload, open(BASELINE, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    print('baseline更新:', BASELINE, '/ allowlist', len(allow), '語')

# ---------------- 短い要約(会話へ返すのはこれだけ) ----------------
for lv in LEVELS:
    ds = daimon_stat[lv]; lr = level_report[lv]
    print(f"{lv}: breadth {lr['cov']}/{lr['tot']}={lr['pct']}% | backlog {lr['backlog']} | "
          f"未紐づけ{ds['unlinked']} | 集中(最大){ds['maxConc']}問/語")
print('級跨ぎ重複(冗長):', len(cross_dup), '語')
print('Excel:', XLSX, '/ sheet:', SHEET)
