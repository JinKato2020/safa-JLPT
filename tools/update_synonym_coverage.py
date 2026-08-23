# 「単語×大問カバー率」シートの【言い換え類義(synonym)】行に「真の母数／真のカバー率」を追記する。
# 全ID母数=級内の全語彙数。真の母数=言い換え可能語(近い類義語があり一意な言い換え問題が作れる語)。
#   分類の正本=src/data/shared/iikaePossible.json(items:{id:{p,syn}})。既存synonym問題を持つ語は定義上p=1。
# 真のカバー率=covered∩possible / 真の母数(=正直なカバー率)。列I=真の母数, 列J=真のカバー率を追加する。
import json, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

GREEN = PatternFill('solid', fgColor='CDE8D4'); YELLOW = PatternFill('solid', fgColor='FCE7C0'); RED = PatternFill('solid', fgColor='F6C9C4')
def signal(p): return GREEN if p >= 80 else (YELLOW if p >= 60 else RED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
V = json.load(open(os.path.join(ROOT, 'src/data/shared/vocab.json'), encoding='utf-8'))
lvlOf = {v['id']: v['level'] for v in V}
vocTot = {lv: sum(1 for v in V if v['level'] == lv) for lv in ('N5', 'N4', 'N3')}

poss = json.load(open(os.path.join(ROOT, 'src/data/shared/iikaePossible.json'), encoding='utf-8'))['items']
trueTot = {lv: 0 for lv in ('N5', 'N4', 'N3')}
for vid, o in poss.items():
    lv = lvlOf.get(vid)
    if lv in trueTot and o.get('p') == 1:
        trueTot[lv] += 1

# 既存synonym問題が対象にしている語(=covered)
covered = {lv: set() for lv in ('N5', 'N4', 'N3')}
for lv in ('N5', 'N4', 'N3'):
    d = json.load(open(os.path.join(ROOT, f'content/problems/moji_goi/synonym_{lv}.json'), encoding='utf-8'))
    for it in d['items']:
        vid = it.get('vocabId')
        if vid and lvlOf.get(vid) == lv:
            covered[lv].add(vid)
trueCov = {lv: len(covered[lv] & {vid for vid, o in poss.items() if o.get('p') == 1 and lvlOf.get(vid) == lv}) for lv in covered}

wb = load_workbook(XLSX)
ws = wb['単語×大問カバー率']
I, J = 9, 10  # 列I=真の母数, 列J=真のカバー率
# ヘッダ行(単語種別で始まる)に列見出しを付ける
for r in range(1, ws.max_row + 1):
    if ws.cell(r, 1).value and str(ws.cell(r, 1).value).startswith('単語種別'):
        ws.cell(r, I, '真の母数'); ws.cell(r, J, '真のカバー率')
        for cc in (I, J):
            ws.cell(r, cc).alignment = Alignment(horizontal='center')

cur_lv = None; done = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and str(a).startswith('■'):
        for lv in ('N5', 'N4', 'N3'):
            if lv in str(a): cur_lv = lv
    b = ws.cell(r, 2).value
    if b and '言い換え類義' in str(b) and cur_lv:
        lv = cur_lv
        cov = len(covered[lv]); allTot = vocTot[lv]
        allPct = round(100 * cov / allTot) if allTot else 0
        tt = trueTot[lv]; tc = trueCov[lv]
        tPct = round(100 * tc / tt) if tt else 0
        ws.cell(r, 3, cov)                 # 問題数(=対象語数)
        ws.cell(r, 4, cov)                 # カバー数
        ws.cell(r, 5, allTot)              # 母数(全ID)
        ws.cell(r, 6, f'{allPct}%'); ws.cell(r, 6).fill = signal(allPct)
        ws.cell(r, I, tt)                  # 真の母数
        jc = ws.cell(r, J, f'{tPct}%'); jc.fill = signal(tPct)
        ws.cell(r, 8, (f'全ID {cov}/{allTot}={allPct}%。真の母数={tt}語(言い換え可能=近い類義語がある語)中 {tc}語={tPct}%。'
                       f'全ID母数には数詞/あいさつ/固有名/類義語のない具体名詞(カメラ等)が含まれ言い換え不可のため、'
                       f'真のカバー率が実力。カタカナ語は類義語がある語のみ計上(ルール→規則等)。分類正本=iikaePossible.json。'))
        done.append((lv, cov, allTot, allPct, tt, tc, tPct))

# 上部の注記も更新(2行目)
ws.cell(2, 1, ('※用法は場面/例文単位でvocabId無し＝語カバー測定不可。漢字は語彙に統合済(漢字カバー行は廃止・2026-08-22)。'
               '言い換え類義は真の母数(言い換え可能語)列を追加(2026-08-23)＝真のカバー率で判断。'))
wb.save(XLSX)
for d in done:
    print(f'{d[0]}: 全ID {d[1]}/{d[2]}={d[3]}%  真 {d[5]}/{d[4]}={d[6]}%')
print('saved', XLSX)
