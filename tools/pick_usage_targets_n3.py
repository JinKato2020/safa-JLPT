# N3 用法の作問対象語を backlog から選び、作問前リスク色で Excel 出力する(セッション直下)。
# 色の意味(作問前スクリーニング=生成後に選択肢単位で再判定する暫定フラグ):
#   RED   = 一意性が壊れやすい語(第2正解リスク)。感情/状態形容詞・近接類義がほぼ同義の語。
#   YELLOW= 誤答づくりが難しい語。漢字が大問級(N3)より上=同音異字ダミー不可/道具不足、または単義の抽象語。
#   なし  = 良ターゲット(自他/選択制限/混同で殺し方を分散しやすい)。
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def L(p): return json.load(open(os.path.join(ROOT, p), encoding='utf-8'))

V = L('src/data/shared/vocab.json')
syn = L('src/data/dict/vocabSynonyms.json')                 # vocabId -> 近接類義語(1語)
kl  = L('src/data/dict/kanjiJlptLevel.json')['items']        # 漢字 -> 級(N5..N1/BEYOND)
try: excluded = set(L('src/data/shared/vocabMetricExcluded.json'))
except Exception: excluded = set()
if isinstance(excluded, dict): excluded = set(excluded)

# covered vocabIds (N4/N3 両大問通算=重複禁止)
cov = set()
for lv in ('N4', 'N3'):
    d = L(f'content/problems/moji_goi/usage_{lv}.json')
    for it in (d['items'] if isinstance(d, dict) else d):
        if it.get('vocabId'): cov.add(it['vocabId'])

RANK = {'N5':0,'N4':1,'N3':2,'N2':3,'N1':4,'BEYOND':5}
KATA = re.compile(r'^[゠-ヿーー]+$')
def kanji_max_level(w):
    ks = [c for c in w if '一' <= c <= '鿿']
    if not ks: return None
    return max((kl.get(c,'BEYOND') for c in ks), key=lambda x: RANK.get(x,5))

def pos_of(v):
    m = (v.get('meaning') or '').strip().lower()
    r = v.get('reading') or ''
    w = v.get('word') or ''
    if m.startswith('to '): return 'verb'
    if KATA.match(r): return 'onomatopoeia'   # 擬音・擬態(すらすら等)は多くカタカナ読み
    # 形容詞: い終わり(word) か na-adj的意味
    if w.endswith('い') and not m.startswith('to '): return 'adj'
    if any(m.startswith(x) for x in ('being ','feeling ','-ly','quietly','suddenly')): return 'adv'
    return 'noun'

EMO = ('happy','sad','angry','afraid','fear','anxious','lonely','painful','glad',
       'pleasant','unpleasant','embarrass','ashamed','jealous','regret','worried',
       'nervous','comfortable','uncomfortable','boring','bored','fun','enjoy',
       'strict','gentle','kind','cruel','beautiful','ugly','cute')

def risk(v):
    pos = pos_of(v)
    m = (v.get('meaning') or '').lower()
    has_syn = v['id'] in syn and (syn[v['id']] or '').strip() != ''
    kmax = kanji_max_level(v['word'])
    above = kmax is not None and RANK.get(kmax,5) > RANK['N3']
    # RED: 一意性リスク
    if pos == 'adj' and (has_syn or any(e in m for e in EMO)):
        return 'RED', '感情/状態形容詞＋近接類義=第2正解リスク(問題ごと落とす候補)'
    if pos == 'adj' and any(e in m for e in EMO):
        return 'RED', '感情形容詞=同義密集'
    # YELLOW: 誤答品質リスク
    if above:
        return 'YELLOW', f'漢字級{kmax}>N3=ルビで同読み化・同音異字ダミー不可/道具不足'
    if pos == 'noun' and not has_syn and len(v['word']) <= 2 and any(a in m for a in ('ness','ity','ism','tion','ment','ance','ence')):
        return 'YELLOW', '単義の抽象語=殺し方を3種に散らしにくい'
    return '', ''

# suitability: verb/onomatopoeia/具体名詞 を上位に。感情形容詞は下位。
def suit(v):
    pos = pos_of(v); s = 0
    s += {'verb':5,'onomatopoeia':4,'noun':3,'adv':3,'adj':1}.get(pos,2)
    if v['id'] in syn: s += 1        # 近接類義がある=用法で外しやすい素材
    r,_ = risk(v)
    if r == 'RED': s -= 4
    if r == 'YELLOW': s -= 1
    return s

import sys
N = int(sys.argv[1]) if len(sys.argv) > 1 else 200   # 抽出語数(既定200・300も可)
back = [v for v in V if v['level']=='N3' and v['id'] not in cov and v['id'] not in excluded]
if os.environ.get('NO_RED') == '1':                 # gen-only時=第2正解リスクのRED語を母集団から除外
    back = [v for v in back if risk(v)[0] != 'RED']
# 本試験に近い品詞バランス(動詞中心＋名詞・副詞/擬態・形容詞)。各品詞内は suitability 順。Nに比例配分。
_BASE = {'verb':96, 'noun':64, 'adv':12, 'onomatopoeia':12, 'adj':16}   # 合計200
QUOTA = {k: round(v * N / 200) for k, v in _BASE.items()}
bypos = {}
for v in back: bypos.setdefault(pos_of(v), []).append(v)
for p in bypos: bypos[p].sort(key=lambda v: (-suit(v), int(v['id'].split('-')[-1])))
pick = []
for p, n in QUOTA.items():
    pick += bypos.get(p, [])[:n]
# 不足分は残り backlog の上位で補充
if len(pick) < N:
    got = {v['id'] for v in pick}
    rest = sorted((v for v in back if v['id'] not in got), key=lambda v:(-suit(v), int(v['id'].split('-')[-1])))
    pick += rest[:N-len(pick)]
pick = pick[:N]
pick.sort(key=lambda v: ({'verb':0,'noun':1,'adv':2,'onomatopoeia':2,'adj':3}.get(pos_of(v),4), -suit(v)))

wb = Workbook(); ws = wb.active; ws.title = f'N3用法_作問対象{N}'
RED = PatternFill('solid', fgColor='F6C9C4'); YEL = PatternFill('solid', fgColor='FCE7C0')
HEAD = PatternFill('solid', fgColor='D9D9D9'); B = Font(bold=True)
cols = ['#','vocabId','語','読み','品詞(推定)','意味(英)','近接類義語','漢字最高級','リスク','リスク理由(作問前・暫定)']
ws.append(cols)
for c in ws[1]: c.fill = HEAD; c.font = B; c.alignment = Alignment(vertical='center')
red_n = yel_n = 0
for i,v in enumerate(pick,1):
    r, why = risk(v)
    kmax = kanji_max_level(v['word']) or '-'
    ws.append([i, v['id'], v['word'], v['reading'], pos_of(v), v.get('meaning',''),
               syn.get(v['id'],''), kmax, {'RED':'一意性','YELLOW':'誤答品質'}.get(r,''), why])
    if r == 'RED':
        red_n += 1
        for cell in ws[ws.max_row]: cell.fill = RED
    elif r == 'YELLOW':
        yel_n += 1
        for cell in ws[ws.max_row]: cell.fill = YEL
widths = [4,10,14,16,12,34,14,10,10,52]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = 'A2'

# 凡例シート
lg = wb.create_sheet('凡例')
for row in [
    ['色','意味','対応'],
    ['赤(一意性)','第2正解が出やすい語(感情/状態形容詞・ほぼ同義の近接類義)','問題ごと落とす/対象語を差し替える候補'],
    ['橙(誤答品質)','誤答を3種の殺し方に散らしにくい語(漢字が大問級より上=同音異字ダミー不可/単義抽象語)','近接類義・選択制限・自他で外す。同音異字型は使わない'],
    ['無色','良ターゲット(自他/選択制限/混同で弁別を測れる)','そのまま作問'],
    ['','',''],
    ['注','この色は「作問前の暫定スクリーニング」。実際の一意性/誤答品質は生成後に選択肢単位で自己監査し、弱い/非一意な誤答セルを本表で赤塗りして再提出する。',''],
]:
    lg.append(row)
lg['A1'].font=B; lg['B1'].font=B; lg['C1'].font=B
for i,w in enumerate([12,60,54],1): lg.column_dimensions[chr(64+i)].width=w

out = os.path.join(ROOT, f'用法N3_作問対象{N}_確認用.xlsx')
wb.save(out)
print('SAVED', out)

# --- 作問ワークフロー用 targets JSON も出力(scratchpad) ---
tgt = []
for v in pick:
    r, why = risk(v)
    tgt.append({'vocabId': v['id'], 'word': v['word'], 'reading': v.get('reading',''),
                'meaning': v.get('meaning',''), 'pos': pos_of(v),
                'synonym': syn.get(v['id'],''), 'kanjiMax': kanji_max_level(v['word']) or '-',
                'risk': r})
tj = os.path.join(ROOT, 'scratchpad', 'usage_n3_300', 'targets.json')
os.makedirs(os.path.dirname(tj), exist_ok=True)
json.dump(tgt, open(tj, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print('TARGETS_JSON', tj)
print('picked', len(pick), 'RED', red_n, 'YELLOW', yel_n, 'clean', len(pick)-red_n-yel_n)
from collections import Counter
print('POS', dict(Counter(pos_of(v) for v in pick)))
print('backlog_total', len(back))
