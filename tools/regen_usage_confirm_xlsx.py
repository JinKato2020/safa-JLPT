# -*- coding: utf-8 -*-
"""用法N4 確認用Excelを「現在のDB(usage_N4.json + タグ)」から作り直す。
apply_usage_new.py はスクラッチパッドの結果ファイルからExcelを作るため、DBを手で直した後は
本ツールで再生成する(= 目視Excelを常にDBと一致させる)。borderline誤答はセルを橙に着色。
使い方: python tools/regen_usage_confirm_xlsx.py [--min 314] [--max 9999]
"""
import json, os, re, sys, argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

JSON_N4 = 'content/problems/moji_goi/usage_N4.json'
SIDE = 'src/data/shared/usageDistractorTags.json'
XLSX = '用法N4_新規作問_確認用.xlsx'
ORANGE = PatternFill('solid', fgColor='FFC000')
HEAD = PatternFill('solid', fgColor='DDDDDD')


def idnum(iid):
    m = re.search(r'(\d+)$', iid)
    return int(m.group(1)) if m else -1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--min', type=int, default=314)
    ap.add_argument('--max', type=int, default=99999)
    a = ap.parse_args()

    items = json.load(open(JSON_N4, encoding='utf-8'))['items']
    tags = json.load(open(SIDE, encoding='utf-8'))['tags']

    rows = []
    for it in items:
        if not it['id'].startswith('N4-V-Y-'):
            continue
        n = idnum(it['id'])
        if not (a.min <= n <= a.max):
            continue
        tg = tags.get(it['id'])
        if not tg:
            continue
        dists = [c for c in it['choices'] if c != it['answer']]
        rows.append((it['id'], it['stem'], it['answer'], dists, tg))
    rows.sort(key=lambda r: idnum(r[0]))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '用法N4 新規作問'
    ws.append(['ID', '語', '○ 正解文', '誤答1', '型1', '誤答2', '型2', '誤答3', '型3'])
    for c in ws[1]:
        c.fill = HEAD; c.font = Font(bold=True)
    ws.append(['凡例: オレンジ=一意性あやしい(要目視)', *[''] * 8])
    bcnt = 0
    for iid, stem, answer, dists, tg in rows:
        line = [iid, stem, answer]
        for j, sent in enumerate(dists):
            t = tg[j] if j < len(tg) else {}
            line += [sent, f"{t.get('type','')}→{t.get('repl','')}"]
        ws.append(line)
        ri = ws.max_row
        for j in range(len(dists)):
            t = tg[j] if j < len(tg) else {}
            if t.get('certainty') == 'borderline':
                ws.cell(ri, 4 + j * 2).fill = ORANGE; bcnt += 1
    for col, w in zip('ABCDEFGHI', [13, 12, 42, 36, 14, 36, 14, 36, 14]):
        ws.column_dimensions[col].width = w
    for r_ in ws.iter_rows():
        for c in r_:
            c.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(XLSX)
    print(f'再生成 {len(rows)}問 / borderline誤答 {bcnt}セル')
    print('XLSX:', os.path.abspath(XLSX))


if __name__ == '__main__':
    main()
