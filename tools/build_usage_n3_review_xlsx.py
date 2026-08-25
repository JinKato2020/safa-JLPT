# 生成した out_*.json を集約し、ユーザーの一意性確認用Excel(セッション直下)を作る。
# 色: 橙=エージェント自己申告 suspect(怪しい誤答) / 赤=自動検出の構造リスク(repl重複=P1違反 / 対象語が文に無い / 同音異字×漢字>N3)。
import json, glob, os, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OD = r'C:\Users\jwpsa\AppData\Local\Temp\claude\c--Users-jwpsa-Documents-desktop-claude-JLPT---\dff2926b-efae-47c7-a4fd-57447b7e8983\scratchpad\usage_n3'
words = {w['vocabId']: w for w in json.load(open(os.path.join(OD, 'words200.json'), encoding='utf-8'))}
RANK = {'N5':0,'N4':1,'N3':2,'N2':3,'N1':4,'BEYOND':5}

items = []
for f in sorted(glob.glob(os.path.join(OD, 'out_*.json'))):
    items += json.load(open(f, encoding='utf-8'))

def stem_core(word):
    # 対象語が文中に(活用形で)出現するかの照合用の不変幹。
    #  ・する動詞(関する/対する) → 漢語部(関/対)。活用しても不変。
    #  ・それ以外(現れる/冷ます/すっぱい) → 末尾1字を落とした形(現れ/冷ま/すっぱ)。活用の直前まで不変。
    #    これで「現す」を使った文(現して=現し)は「現れ」を含まず=誤答文に対象語なしを正しく検出する。
    if len(word) <= 1:
        return [word]
    if word.endswith('する') and len(word) > 2:
        return [word[:-2]]
    return [word[:-1]]

ORANGE = PatternFill('solid', fgColor='FCE7C0')
RED = PatternFill('solid', fgColor='F6C9C4')
GREEN = PatternFill('solid', fgColor='CDE8D4')
HEAD = PatternFill('solid', fgColor='D9D9D9'); B = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook(); ws = wb.active; ws.title = 'N3用法_200問'
cols = ['#','vocabId','語','読み','○正用文',
        '✕誤答1','repl1','型1','✕誤答2','repl2','型2','✕誤答3','repl3','型3','自動チェック']
ws.append(cols)
for c in ws[1]: c.fill = HEAD; c.font = B
ws['E1'].fill = GREEN

# 誤答セルの列インデックス(1始まり): 誤答1=6, 誤答2=9, 誤答3=12
DCOL = [6, 9, 12]
n_suspect = n_red = 0
for i, q in enumerate(items, 1):
    w = words.get(q['vocabId'], {})
    kmax = w.get('kanjiMax'); above = RANK.get(kmax, 2) > RANK['N3']
    ds = q.get('distractors', [])[:3]
    while len(ds) < 3: ds.append({'sentence':'(欠落)','repl':'','type':'','suspect':True,'reason':'誤答不足'})
    row = [i, q['vocabId'], q.get('word',''), q.get('reading',''), q.get('correct','')]
    for d in ds:
        row += [d.get('sentence',''), d.get('repl',''), d.get('type','')]
    # 自動チェック
    notes = []
    repls = [d.get('repl','') for d in ds]
    dup = [r for r in set(repls) if r and repls.count(r) > 1]
    if dup: notes.append(f'P1違反:repl重複[{"/".join(dup)}]')
    types = [d.get('type','') for d in ds]
    if len(set(t for t in types if t)) < 2: notes.append('P2注意:型が単一')
    # 対象語の出現チェック(正用+各誤答)
    cores = stem_core(q.get('word',''))
    if not any(c in (q.get('correct','') or '') for c in cores):
        notes.append('正用文に対象語なし?')
    row.append('')  # placeholder for notes col (index 15)
    ws.append(row)
    r = ws.max_row
    for c in ws[r]: c.alignment = WRAP
    ws.cell(r, 5).fill = GREEN
    # 誤答ごとの色付け
    for j, d in enumerate(ds):
        col = DCOL[j]
        red = False
        if d.get('repl','') in dup:
            red = True
        if not any(c in (d.get('sentence','') or '') for c in cores):
            red = True; notes.append(f'誤答{j+1}に対象語なし?')
        if above and ('同音' in d.get('type','') or 'かな固定' in d.get('type','') or '多義' in d.get('type','')):
            red = True; notes.append(f'誤答{j+1}:漢字{kmax}>N3で同読み化リスク')
        if red:
            ws.cell(r, col).fill = RED; n_red += 1
        elif d.get('suspect'):
            ws.cell(r, col).fill = ORANGE; n_suspect += 1
            if d.get('reason'): notes.append(f'誤答{j+1}(自己申告):{d["reason"]}')
        elif d.get('suspect'):
            pass
    ws.cell(r, 15).value = ' / '.join(notes)
    if any(('違反' in x or 'なし' in x or '>N3' in x) for x in notes):
        ws.cell(r, 15).fill = RED
    elif any('自己申告' in x for x in notes):
        ws.cell(r, 15).fill = ORANGE
    # P2注意など設計メモのみの行はセルを塗らない(色の意味を「怪しい誤答/構造欠陥」に限定)

widths = [4,10,12,12,40,34,9,12,34,9,12,34,9,12,30]
for i, wd in enumerate(widths, 1): ws.column_dimensions[chr(64+i)].width = wd
ws.freeze_panes = 'A2'

lg = wb.create_sheet('凡例')
for row in [
    ['色','意味','あなたの作業'],
    ['緑','○ 正しい用法の文(正解)','—'],
    ['橙','エージェントが「怪しい」と自己申告した誤答(第2正解の疑い)','一意性を重点確認'],
    ['赤','自動検出の構造リスク: repl重複(P1違反)/対象語が文に無い/同音異字×漢字>N3','要修正の可能性大'],
    ['','',''],
    ['注','色の無い誤答も一意性は未保証。最終確認はあなたが全問で行う前提。','']]:
    lg.append(row)
for c in lg[1]: c.font = B
for i, wd in enumerate([8,58,40], 1): lg.column_dimensions[chr(64+i)].width = wd

out = os.path.join(ROOT, '用法N3_作問_確認用.xlsx')
wb.save(out)
print('SAVED', out)
print('questions', len(items), 'orange_cells', n_suspect, 'red_cells', n_red)
