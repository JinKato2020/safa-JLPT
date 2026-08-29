# -*- coding: utf-8 -*-
"""新規N3用法300問(N3-V-Y-0352..0651)の一意性・目視確認用Excelをセッション直下に作る。
DB(usage_N3.json)＋タグ＋targets(risk/読み)から再生成。
色: 橙=作問前YELLOWリスク(漢字級>N3=同音異字ダミー不可等)/黄=monoType単一型。ユーザーが○×を目視確認する。
使い方: python tools/build_usage_n3_review_300.py
"""
import json, os, re, sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
L = lambda p: json.load(open(os.path.join(ROOT, p), encoding='utf-8'))
strip = lambda s: re.sub(r'（[^）]*）', '', s or '')

d = L('content/problems/moji_goi/usage_N3.json')['items']
TAG = L('src/data/shared/usageDistractorTags.json')
tags = TAG['tags']; mono = set(TAG.get('monoTypeAllow', []))
tgt = {t['vocabId']: t for t in L('scratchpad/usage_n3_300/targets.json')}

def idnum(i):
    m = re.search(r'(\d+)$', i); return int(m.group(1)) if m else -1

new = [it for it in d if it.get('id', '').startswith('N3-V-Y') and 352 <= idnum(it['id']) <= 651]
new.sort(key=lambda it: idnum(it['id']))

ORANGE = PatternFill('solid', fgColor='FCE7C0')
YELLOW = PatternFill('solid', fgColor='FFF2B2')
HEAD = PatternFill('solid', fgColor='D9D9D9'); GREEN = PatternFill('solid', fgColor='CDE8D4')
B = Font(bold=True); WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook(); ws = wb.active; ws.title = 'N3用法_新規300'
cols = ['#', 'id', 'vocabId', '語', '読み', 'pos', 'risk', '○正用文',
        '✕誤答1', 'repl1', '型1', '✕誤答2', 'repl2', '型2', '✕誤答3', 'repl3', '型3', 'mono']
ws.append(cols)
for c in ws[1]: c.fill = HEAD; c.font = B
ws['H1'].fill = GREEN

for n, it in enumerate(new, 1):
    vid = it['vocabId']; t = tgt.get(vid, {})
    tg = tags.get(it['id'], [{}, {}, {}])
    ans = strip(it['answer'])
    drs = [strip(c) for c in it['choices'][1:4]]
    row = [n, it['id'], vid, it.get('stem', ''), t.get('reading', ''), t.get('pos', ''),
           t.get('risk', ''), ans,
           drs[0], tg[0].get('repl', ''), tg[0].get('type', ''),
           drs[1], tg[1].get('repl', ''), tg[1].get('type', ''),
           drs[2], tg[2].get('repl', ''), tg[2].get('type', ''),
           'mono' if it['id'] in mono else '']
    ws.append(row)
    r = ws.max_row
    for c in ws[r]: c.alignment = WRAP
    if t.get('risk') == 'YELLOW':
        for col in ('D', 'G'): ws[f'{col}{r}'].fill = ORANGE
    if it['id'] in mono:
        ws[f'R{r}'].fill = YELLOW

widths = [4, 12, 10, 10, 12, 6, 8, 40, 34, 8, 6, 34, 8, 6, 34, 8, 6, 6]
for i, w in enumerate(widths, 1): ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = 'A2'

lg = wb.create_sheet('凡例')
for r in [['色', '意味'], ['橙(語/risk列)', '作問前YELLOW=漢字級>N3等。同音異字ダミー不可ゆえ近接/選択/自他で作成。目視で一意性を確認'],
          ['黄(mono列)', '3誤答が単一の殺し方(選択制限型など)。公式が認める良問なら可'],
          ['緑(見出し)', '正用文=対象語を正しく使った唯一の文'],
          ['', ''], ['確認観点', '各誤答は対象語だと明確に不自然か / replを入れれば自然か / 対象語も成立する第2正解が無いか']]:
    lg.append(r)
for i, w in enumerate([16, 84], 1): lg.column_dimensions[chr(64 + i)].width = w
lg['A1'].font = B; lg['B1'].font = B

out = os.path.join(ROOT, '用法N3_新規300_確認用.xlsx')
wb.save(out)
print('SAVED', out)
print('rows', len(new), 'YELLOW', sum(1 for it in new if tgt.get(it['vocabId'], {}).get('risk') == 'YELLOW'),
      'mono', sum(1 for it in new if it['id'] in mono))
