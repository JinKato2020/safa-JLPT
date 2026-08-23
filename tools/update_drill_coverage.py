# 「学習ドリル×カバー率」シートを在庫Excelへ生成/更新する。
# 数値の正本 = tools/drill_coverage.ts（アプリ本体の eligible 関数を直接呼ぶ）が書く
#   scratchpad/pg/drill_coverage.json。先に `node --import tsx tools/drill_coverage.ts` を実行すること。
# 学習タブの各ドリルが「対象にできる項目(ID紐づけ)」の広さ＝カバー率。試験タブ側は別シート「単語×大問カバー率」。
import json, os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

GREEN = PatternFill('solid', fgColor='CDE8D4')   # ≥80%
YELLOW = PatternFill('solid', fgColor='FCE7C0')   # 60-79%
RED = PatternFill('solid', fgColor='F6C9C4')      # <60%
HDR = PatternFill('solid', fgColor='2E5A88')

def signal(p):
    return GREEN if p >= 80 else (YELLOW if p >= 60 else RED)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'memory', '在庫・模試ストックまとめ.xlsx')
DATA = os.path.join(ROOT, 'scratchpad', 'pg', 'drill_coverage.json')

rows = json.load(open(DATA, encoding='utf-8'))

# 各ドリルの低カバーの理由(点検メモ)
NOTE = {
    '語彙パズル(産出)': 'かなタイルで単語を作る産出ドリル。カタカナ語・波ダッシュ語(～)・2〜6モーラ外の語は対象外ゆえ100%未満(欠落ではなく仕様)。',
    '聞き取り(語彙)': '読みを聞いて単語を選ぶ。非自立語(接辞/助詞)・波ダッシュ語のみ対象外。',
    '意味を選ぶ(受容)': '文法点の意味を4択。意味を持つ全点が対象。',
    '文法パズル(産出)': '例文の空所に文法語をタイルで作る。例文に表層形が一意に1回出る点のみ対象(0回/複数回は空所位置が曖昧で除外)＝構造的に低い。上げたい場合は例文を「表層形が1回だけ出る」形に整えるか、対象点の例文を追加する。',
    '聞き取り(漢字)': '読みを聞いて漢字を選ぶ。drillレップを持つ当該級の漢字が対象。',
}

wb = load_workbook(XLSX)
SHEET = '学習ドリル×カバー率'
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 20
for c in ('C', 'D', 'E', 'F', 'G'):
    ws.column_dimensions[c].width = 11
ws.column_dimensions['H'].width = 60

r = 1
ws.cell(r, 1, '学習ドリル×カバー率（学習タブ・各ドリルがID紐づけで対象にできる項目の広さ）')
ws.cell(r, 1).font = Font(bold=True)
r += 1
ws.cell(r, 1, '※全ID母数=レベル内の全項目(語彙数/文法点数[n5-g-92除外]/漢字数)。真の母数=そのドリルで「そもそも作れる」項目数(カタカナ語・分離型文法など"永久に不可能"を除く)＝真のカバー率が実力。試験タブの大問は別シート「単語×大問カバー率」。')
r += 2

HDRS = ['区分', 'ドリル', '対象', '全ID母数', '全IDカバー率', '真の母数', '真のカバー率', '点検メモ']
for lv in ('N5', 'N4', 'N3'):
    ws.cell(r, 1, f'■ {lv}'); ws.cell(r, 1).font = Font(bold=True)
    r += 1
    for i, h in enumerate(HDRS, start=1):
        cell = ws.cell(r, i, h)
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = HDR
        cell.alignment = Alignment(horizontal='center', vertical='center')
    r += 1
    for row in rows[lv]:
        cov, tot, ach = row['cov'], row['tot'], row['achievable']
        p_all = round(cov / tot * 100) if tot else 0
        p_true = min(100, round(cov / ach * 100)) if ach else 0
        ws.cell(r, 1, row['kubun'])
        ws.cell(r, 2, row['drill'])
        ws.cell(r, 3, cov)
        ws.cell(r, 4, tot)
        c_all = ws.cell(r, 5, f'{p_all}%'); c_all.fill = signal(p_all)
        ws.cell(r, 6, ach)
        c_true = ws.cell(r, 7, f'{p_true}%'); c_true.fill = signal(p_true)
        note = NOTE.get(row['drill'], '')
        if p_all < p_true:
            note = f'見かけ{p_all}%だが真の{p_true}%＝差{p_true-p_all}ptは"作れない天井"。' + note
        m = ws.cell(r, 8, note); m.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1
    r += 1

ws.cell(r, 1, '凡例：緑=問題なし(≥80%) ／ 黄=気になる(60-79%) ／ 赤=危険(<60%)。カバー率が2つあるのは"天井"のあるドリル＝真のカバー率で判断。')
wb.save(XLSX)
print('OK シート更新:', SHEET, '→', XLSX)
